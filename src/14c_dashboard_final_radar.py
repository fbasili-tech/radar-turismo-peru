from pathlib import Path
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabelList


# ============================================================
# RADAR TURISMO DE NATURALEZA Y AVENTURA - PERÚ
# ETAPA 14C: DASHBOARD FINAL CORREGIDO
# ============================================================

ENTRADA = Path("outputs/radar_turismo_tablero_base_2026.xlsx")
SALIDA = Path("outputs/radar_turismo_dashboard_final_2026.xlsx")


# ============================================================
# COLORES
# ============================================================

VERDE_OSCURO = "143D33"
VERDE = "2E7D5B"
AZUL = "2F6B9A"
ROJO = "D9534F"
NARANJA = "F0A04B"
AMARILLO = "E4C441"
GRIS_CLARO = "F3F5F7"
GRIS_OSCURO = "404040"
BLANCO = "FFFFFF"


# ============================================================
# FUNCIONES
# ============================================================

def color_alerta(valor):

    mapa = {
        "ROJA": ROJO,
        "NARANJA": NARANJA,
        "AMARILLA": AMARILLO,
        "VERDE": VERDE
    }

    return mapa.get(
        str(valor).upper(),
        "D9E1E8"
    )


def bloque_kpi(ws, rango, titulo, valor, color):

    ws.merge_cells(rango)

    celda = ws[
        rango.split(":")[0]
    ]

    celda.value = f"{titulo}\n{valor}"

    celda.fill = PatternFill(
        "solid",
        fgColor=color
    )

    celda.font = Font(
        size=14,
        bold=True,
        color=BLANCO
    )

    celda.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )


# ============================================================
# PROCESO
# ============================================================

def main():

    print("=" * 100)
    print("RADAR TURISMO NATURALEZA Y AVENTURA - PERÚ")
    print("14C - DASHBOARD FINAL CORREGIDO")
    print("=" * 100)

    # --------------------------------------------------------
    # CARGAR DATOS
    # --------------------------------------------------------

    ranking = pd.read_excel(
        ENTRADA,
        sheet_name="Ranking_Nacional"
    )

    alertas = pd.read_excel(
        ENTRADA,
        sheet_name="Alertas_Territoriales"
    )

    oportunidades = pd.read_excel(
        ENTRADA,
        sheet_name="Oportunidades"
    )

    resumen = pd.read_excel(
        ENTRADA,
        sheet_name="Resumen_Nacional"
    )

    print(
        f"\nTerritorios cargados: {len(ranking)}"
    )

    ranking = (
        ranking
        .sort_values(
            "IRNA_Estructural",
            ascending=False
        )
        .reset_index(drop=True)
    )

    mediana_estructural = ranking[
        "IRNA_Estructural"
    ].median()

    mediana_ejecucion = ranking[
        "IRNA_Ejecucion"
    ].median()

    # --------------------------------------------------------
    # CREAR ARCHIVO
    # --------------------------------------------------------

    with pd.ExcelWriter(
        SALIDA,
        engine="openpyxl"
    ) as writer:

        ranking.to_excel(
            writer,
            sheet_name="Ranking",
            index=False
        )

        alertas.to_excel(
            writer,
            sheet_name="Alertas",
            index=False
        )

        oportunidades.to_excel(
            writer,
            sheet_name="Oportunidades",
            index=False
        )

        resumen.to_excel(
            writer,
            sheet_name="Indicadores",
            index=False
        )

        pd.DataFrame().to_excel(
            writer,
            sheet_name="Dashboard",
            index=False
        )

    wb = load_workbook(SALIDA)

    ws = wb["Dashboard"]

    # Dashboard primero
    wb._sheets.remove(ws)
    wb._sheets.insert(0, ws)

    ws.sheet_view.showGridLines = False

    # ========================================================
    # TÍTULO
    # ========================================================

    ws.merge_cells("A1:P2")

    ws["A1"] = (
        "RADAR DE TURISMO DE NATURALEZA Y AVENTURA DEL PERÚ"
    )

    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=VERDE_OSCURO
    )

    ws["A1"].font = Font(
        size=22,
        bold=True,
        color=BLANCO
    )

    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    ws.merge_cells("A3:P3")

    ws["A3"] = (
        "Inversión pública, fortaleza territorial, "
        "ejecución y brechas de gestión | 2026"
    )

    ws["A3"].font = Font(
        size=11,
        italic=True,
        color=GRIS_OSCURO
    )

    ws["A3"].alignment = Alignment(
        horizontal="center"
    )

    # ========================================================
    # KPIs
    # ========================================================

    pim_total = ranking[
        "PIM_Radar"
    ].sum()

    lider = ranking.iloc[0]

    consolidado_df = ranking[
        ranking["Estado_Territorial"]
        == "FORTALEZA CONSOLIDADA"
    ]

    if len(consolidado_df) > 0:
        consolidado = consolidado_df.iloc[0][
            "Departamento"
        ]
    else:
        consolidado = "—"

    alertas_rojas = (
        ranking["Nivel_Alerta"] == "ROJA"
    ).sum()

    bloque_kpi(
        ws,
        "A5:C8",
        "TERRITORIOS ANALIZADOS",
        len(ranking),
        AZUL
    )

    bloque_kpi(
        ws,
        "D5:F8",
        "PIM NÚCLEO RADAR",
        f"S/ {pim_total / 1_000_000:.1f} M",
        VERDE
    )

    bloque_kpi(
        ws,
        "G5:I8",
        "LÍDER ESTRUCTURAL",
        lider["Departamento"],
        VERDE_OSCURO
    )

    bloque_kpi(
        ws,
        "J5:L8",
        "FORTALEZA CONSOLIDADA",
        consolidado,
        AZUL
    )

    bloque_kpi(
        ws,
        "M5:P8",
        "ALERTAS ROJAS",
        int(alertas_rojas),
        ROJO
    )

    # ========================================================
    # TOP 10 IRNA
    # ========================================================

    ws["A10"] = "TOP 10 — IRNA ESTRUCTURAL"

    ws["A10"].font = Font(
        size=14,
        bold=True,
        color=VERDE_OSCURO
    )

    encabezados = [
        "Departamento",
        "IRNA",
        "Ejecución",
        "Brecha",
        "Alerta"
    ]

    for col, texto in enumerate(
        encabezados,
        start=1
    ):

        c = ws.cell(
            11,
            col,
            texto
        )

        c.fill = PatternFill(
            "solid",
            fgColor=VERDE_OSCURO
        )

        c.font = Font(
            bold=True,
            color=BLANCO
        )

        c.alignment = Alignment(
            horizontal="center"
        )

    top10 = ranking.head(10)

    for i, (_, row) in enumerate(
        top10.iterrows(),
        start=12
    ):

        ws.cell(
            i,
            1,
            row["Departamento"]
        )

        ws.cell(
            i,
            2,
            row["IRNA_Estructural"]
        )

        ws.cell(
            i,
            3,
            row["IRNA_Ejecucion"]
        )

        ws.cell(
            i,
            4,
            row["Brecha_IRNA"]
        )

        alerta = ws.cell(
            i,
            5,
            row["Nivel_Alerta"]
        )

        alerta.fill = PatternFill(
            "solid",
            fgColor=color_alerta(
                row["Nivel_Alerta"]
            )
        )

        alerta.font = Font(
            bold=True,
            color=BLANCO
        )

        alerta.alignment = Alignment(
            horizontal="center"
        )

    # ========================================================
    # GRÁFICO TOP 10
    # UNA SOLA SERIE
    # SIN LEYENDA
    # ========================================================

    chart_top = BarChart()

    chart_top.type = "bar"
    chart_top.style = 10

    chart_top.title = (
        "Top 10 IRNA estructural"
    )

    chart_top.x_axis.title = "IRNA"

    # No necesitamos título del eje vertical
    chart_top.y_axis.title = None

    valores = Reference(
        ws,
        min_col=2,
        min_row=12,
        max_row=21
    )

    categorias = Reference(
        ws,
        min_col=1,
        min_row=12,
        max_row=21
    )

    chart_top.add_data(
        valores,
        titles_from_data=False,
        from_rows=False
    )

    chart_top.set_categories(
        categorias
    )

    # Eliminar leyenda
    chart_top.legend = None

    # Áncash arriba
    chart_top.y_axis.reverseOrder = True

    chart_top.height = 9
    chart_top.width = 18

    # Solo mostrar valores
    chart_top.dataLabels = DataLabelList()

    chart_top.dataLabels.showVal = True
    chart_top.dataLabels.showLegendKey = False
    chart_top.dataLabels.showCatName = False
    chart_top.dataLabels.showSerName = False
    chart_top.dataLabels.showPercent = False

    # Una sola apariencia
    chart_top.series[
        0
    ].graphicalProperties.solidFill = VERDE

    chart_top.series[
        0
    ].graphicalProperties.line.solidFill = VERDE

    ws.add_chart(
        chart_top,
        "G10"
    )

    # ========================================================
    # TOP BRECHAS
    # ========================================================

    ws["A24"] = (
        "TOP BRECHAS ENTRE POTENCIAL Y EJECUCIÓN"
    )

    ws["A24"].font = Font(
        size=14,
        bold=True,
        color=ROJO
    )

    encabezados_b = [
        "Departamento",
        "IRNA Estructural",
        "Ejecución",
        "Brecha"
    ]

    for col, texto in enumerate(
        encabezados_b,
        start=1
    ):

        c = ws.cell(
            25,
            col,
            texto
        )

        c.fill = PatternFill(
            "solid",
            fgColor=ROJO
        )

        c.font = Font(
            bold=True,
            color=BLANCO
        )

    top_brechas = (
        ranking
        .sort_values(
            "Brecha_IRNA",
            ascending=False
        )
        .head(8)
    )

    for i, (_, row) in enumerate(
        top_brechas.iterrows(),
        start=26
    ):

        ws.cell(
            i,
            1,
            row["Departamento"]
        )

        ws.cell(
            i,
            2,
            row["IRNA_Estructural"]
        )

        ws.cell(
            i,
            3,
            row["IRNA_Ejecucion"]
        )

        ws.cell(
            i,
            4,
            row["Brecha_IRNA"]
        )

    # ========================================================
    # GRÁFICO BRECHAS
    # UNA SOLA SERIE
    # ========================================================

    chart_b = BarChart()

    chart_b.type = "bar"
    chart_b.style = 10

    chart_b.title = (
        "Principales brechas territoriales"
    )

    chart_b.x_axis.title = "Brecha IRNA"
    chart_b.y_axis.title = None

    valores_b = Reference(
        ws,
        min_col=4,
        min_row=26,
        max_row=33
    )

    categorias_b = Reference(
        ws,
        min_col=1,
        min_row=26,
        max_row=33
    )

    chart_b.add_data(
        valores_b,
        titles_from_data=False,
        from_rows=False
    )

    chart_b.set_categories(
        categorias_b
    )

    chart_b.legend = None

    # Mayor brecha arriba
    chart_b.y_axis.reverseOrder = True

    chart_b.height = 8
    chart_b.width = 18

    chart_b.dataLabels = DataLabelList()

    chart_b.dataLabels.showVal = True
    chart_b.dataLabels.showLegendKey = False
    chart_b.dataLabels.showCatName = False
    chart_b.dataLabels.showSerName = False
    chart_b.dataLabels.showPercent = False

    chart_b.series[
        0
    ].graphicalProperties.solidFill = ROJO

    chart_b.series[
        0
    ].graphicalProperties.line.solidFill = ROJO

    ws.add_chart(
        chart_b,
        "G25"
    )

    # ========================================================
    # MATRIZ POTENCIAL VS EJECUCIÓN
    # ========================================================

    ws["A36"] = (
        "POTENCIAL ESTRUCTURAL VS EJECUCIÓN"
    )

    ws["A36"].font = Font(
        size=14,
        bold=True,
        color=AZUL
    )

    # Datos auxiliares ocultos
    fila_aux = 70

    ws.cell(
        fila_aux,
        1,
        "Departamento"
    )

    ws.cell(
        fila_aux,
        2,
        "IRNA"
    )

    ws.cell(
        fila_aux,
        3,
        "Ejecución"
    )

    for i, row in ranking.iterrows():

        fila = fila_aux + i + 1

        ws.cell(
            fila,
            1,
            row["Departamento"]
        )

        ws.cell(
            fila,
            2,
            row["IRNA_Estructural"]
        )

        ws.cell(
            fila,
            3,
            row["IRNA_Ejecucion"]
        )

    scatter = ScatterChart()

    scatter.title = (
        "IRNA estructural vs ejecución"
    )

    scatter.x_axis.title = (
        "IRNA Estructural"
    )

    scatter.y_axis.title = (
        "Ejecución presupuestal (%)"
    )

    scatter.height = 11
    scatter.width = 17

    xvalues = Reference(
        ws,
        min_col=2,
        min_row=fila_aux + 1,
        max_row=fila_aux + len(ranking)
    )

    yvalues = Reference(
        ws,
        min_col=3,
        min_row=fila_aux + 1,
        max_row=fila_aux + len(ranking)
    )

    serie = Series(
        yvalues,
        xvalues,
        title="Territorios"
    )

    # Puntos sin líneas
    serie.graphicalProperties.line.noFill = True

    serie.marker.symbol = "circle"
    serie.marker.size = 8

    serie.marker.graphicalProperties.solidFill = AZUL
    serie.marker.graphicalProperties.line.solidFill = AZUL

    scatter.series.append(
        serie
    )

    scatter.legend = None

    ws.add_chart(
        scatter,
        "A38"
    )

    # ========================================================
    # REFERENCIAS
    # ========================================================

    ws["J36"] = "REFERENCIAS NACIONALES"

    ws["J36"].font = Font(
        bold=True,
        color=AZUL
    )

    ws["J37"] = (
        f"Mediana IRNA estructural: "
        f"{mediana_estructural:.1f}"
    )

    ws["J38"] = (
        f"Mediana ejecución: "
        f"{mediana_ejecucion:.1f}%"
    )

    # ========================================================
    # LECTURA ESTRATÉGICA
    # ========================================================

    mejores_ejecucion = (
        ranking
        .sort_values(
            "IRNA_Ejecucion",
            ascending=False
        )
        .head(2)["Departamento"]
        .tolist()
    )

    ws.merge_cells(
        "J40:P53"
    )

    ws["J40"] = (
        "LECTURA ESTRATÉGICA\n\n"

        f"• {lider['Departamento']} lidera la fortaleza "
        f"estructural con IRNA "
        f"{lider['IRNA_Estructural']:.1f}, pero registra "
        f"una brecha de {lider['Brecha_IRNA']:.1f} puntos.\n\n"

        f"• {consolidado} aparece como la principal "
        f"fortaleza consolidada del Radar.\n\n"

        f"• {mejores_ejecucion[0]} y "
        f"{mejores_ejecucion[1]} presentan los mayores "
        f"niveles de ejecución del núcleo Radar.\n\n"

        f"• {alertas_rojas} territorios se encuentran "
        f"en alerta roja por brechas críticas de gestión."
    )

    ws["J40"].fill = PatternFill(
        "solid",
        fgColor=GRIS_CLARO
    )

    ws["J40"].font = Font(
        size=11,
        color=GRIS_OSCURO
    )

    ws["J40"].alignment = Alignment(
        vertical="top",
        wrap_text=True
    )

    # ========================================================
    # NOTA METODOLÓGICA
    # ========================================================

    ws.merge_cells(
        "A56:P61"
    )

    ws["A56"] = (
        "NOTA METODOLÓGICA\n"
        "IRNA Estructural integra inversión, vocación "
        "territorial, masa crítica, diversidad y evidencia. "
        "IRNA Ejecución refleja el avance presupuestal. "
        "La Brecha IRNA expresa la diferencia entre "
        "fortaleza estructural y ejecución. "
        f"Las medianas nacionales son "
        f"{mediana_estructural:.1f} para IRNA estructural "
        f"y {mediana_ejecucion:.1f}% para ejecución."
    )

    ws["A56"].fill = PatternFill(
        "solid",
        fgColor=GRIS_CLARO
    )

    ws["A56"].font = Font(
        size=10,
        color=GRIS_OSCURO
    )

    ws["A56"].alignment = Alignment(
        vertical="top",
        wrap_text=True
    )

    # ========================================================
    # DIMENSIONES
    # ========================================================

    widths = {
        "A": 23,
        "B": 14,
        "C": 14,
        "D": 14,
        "E": 12,
        "F": 3,
        "G": 14,
        "H": 14,
        "I": 14,
        "J": 16,
        "K": 16,
        "L": 16,
        "M": 16,
        "N": 16,
        "O": 16,
        "P": 16
    }

    for col, width in widths.items():

        ws.column_dimensions[
            col
        ].width = width

    for fila in range(
        1,
        62
    ):

        ws.row_dimensions[
            fila
        ].height = 22

    ws.row_dimensions[
        1
    ].height = 32

    # Ocultar datos auxiliares
    for fila in range(
        fila_aux,
        fila_aux + len(ranking) + 2
    ):

        ws.row_dimensions[
            fila
        ].hidden = True

    # ========================================================
    # GUARDAR
    # ========================================================

    wb.save(SALIDA)

    print("\n" + "=" * 100)
    print("DASHBOARD FINAL CORREGIDO")
    print("=" * 100)

    print(SALIDA)

    print(
        "\n✓ DASHBOARD FINAL DEL RADAR COMPLETADO"
    )

    print("FIN")


if __name__ == "__main__":
    main()