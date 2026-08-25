from pathlib import Path
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)
from openpyxl.chart import (
    BarChart,
    ScatterChart,
    Reference,
    Series
)
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter


# ============================================================
# RADAR TURISMO DE NATURALEZA Y AVENTURA - PERÚ
# ETAPA 14: DASHBOARD EJECUTIVO EXCEL
# ============================================================

ENTRADA = Path(
    "outputs/radar_turismo_tablero_base_2026.xlsx"
)

SALIDA = Path(
    "outputs/radar_turismo_dashboard_ejecutivo_2026.xlsx"
)


# ============================================================
# PALETA
# ============================================================

VERDE_OSCURO = "173F35"
VERDE = "287A5A"
VERDE_CLARO = "DDEFE7"

AZUL = "235789"
AZUL_CLARO = "DDEBF7"

NARANJA = "F4A261"
AMARILLO = "F4D35E"
ROJO = "D9534F"

GRIS_OSCURO = "404040"
GRIS = "D9E1E8"
GRIS_CLARO = "F3F5F7"

BLANCO = "FFFFFF"


# ============================================================
# FUNCIONES
# ============================================================

def ajustar_columnas(ws, limites=None):

    if limites is None:
        limites = {}

    for columna in ws.columns:

        letra = get_column_letter(
            columna[0].column
        )

        max_length = 0

        for celda in columna:

            try:
                if celda.value is not None:
                    max_length = max(
                        max_length,
                        len(str(celda.value))
                    )
            except Exception:
                pass

        ancho = min(
            max(max_length + 2, 10),
            limites.get(letra, 35)
        )

        ws.column_dimensions[
            letra
        ].width = ancho


def estilo_encabezado(ws, fila=1):

    fill = PatternFill(
        "solid",
        fgColor=VERDE_OSCURO
    )

    font = Font(
        color=BLANCO,
        bold=True
    )

    for cell in ws[fila]:

        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )


def formato_moneda(celda):

    celda.number_format = (
        '"S/ " #,##0'
    )


def formato_porcentaje(celda):

    celda.number_format = '0.0"%"'


def color_alerta(valor):

    mapa = {
        "ROJA": ROJO,
        "NARANJA": NARANJA,
        "AMARILLA": AMARILLO,
        "VERDE": VERDE
    }

    return mapa.get(
        str(valor).upper(),
        GRIS
    )


# ============================================================
# PROCESO PRINCIPAL
# ============================================================

def main():

    print("=" * 100)
    print("RADAR TURISMO NATURALEZA Y AVENTURA - PERÚ")
    print("14 - DASHBOARD EJECUTIVO EXCEL")
    print("=" * 100)

    # --------------------------------------------------------
    # CARGAR DATOS
    # --------------------------------------------------------

    ranking = pd.read_excel(
        ENTRADA,
        sheet_name="Ranking_Nacional"
    )

    matriz = pd.read_excel(
        ENTRADA,
        sheet_name="Potencial_vs_Ejecucion"
    )

    alertas = pd.read_excel(
        ENTRADA,
        sheet_name="Alertas_Territoriales"
    )

    oportunidades = pd.read_excel(
        ENTRADA,
        sheet_name="Oportunidades"
    )

    resumen = pd.read_excel(
        ENTRADA,
        sheet_name="Resumen_Nacional"
    )

    print(
        f"\nTerritorios cargados: {len(ranking)}"
    )

    # --------------------------------------------------------
    # ETIQUETAS DE CUADRANTES MÁS PRECISAS
    # --------------------------------------------------------

    matriz["Cuadrante_IRNA"] = (
        matriz["Cuadrante_IRNA"]
        .replace(
            {
                "Q1 - ALTO POTENCIAL / ALTA EJECUCIÓN":
                    "Q1 - SOBRE MEDIANA EN POTENCIAL Y EJECUCIÓN",

                "Q2 - ALTO POTENCIAL / BAJA EJECUCIÓN":
                    "Q2 - SOBRE MEDIANA EN POTENCIAL / BAJO MEDIANA EN EJECUCIÓN",

                "Q3 - MENOR POTENCIAL / ALTA EJECUCIÓN":
                    "Q3 - BAJO MEDIANA EN POTENCIAL / SOBRE MEDIANA EN EJECUCIÓN",

                "Q4 - MENOR POTENCIAL / BAJA EJECUCIÓN":
                    "Q4 - BAJO MEDIANA EN POTENCIAL Y EJECUCIÓN"
            }
        )
    )

    # --------------------------------------------------------
    # EXPORTAR BASE PRIMERO
    # --------------------------------------------------------

    with pd.ExcelWriter(
        SALIDA,
        engine="openpyxl"
    ) as writer:

        ranking.to_excel(
            writer,
            sheet_name="Ranking",
            index=False
        )

        matriz.to_excel(
            writer,
            sheet_name="Potencial_Ejecucion",
            index=False
        )

        alertas.to_excel(
            writer,
            sheet_name="Alertas",
            index=False
        )

        oportunidades.to_excel(
            writer,
            sheet_name="Oportunidades",
            index=False
        )

        resumen.to_excel(
            writer,
            sheet_name="Indicadores",
            index=False
        )

        # Crear inicialmente el Dashboard vacío
        pd.DataFrame().to_excel(
            writer,
            sheet_name="Dashboard",
            index=False
        )

    # ========================================================
    # FORMATO CON OPENPYXL
    # ========================================================

    wb = load_workbook(
        SALIDA
    )

    # Mover dashboard al inicio
    ws_dash = wb["Dashboard"]

    wb._sheets.remove(
        ws_dash
    )

    wb._sheets.insert(
        0,
        ws_dash
    )

    # ========================================================
    # DASHBOARD
    # ========================================================

    ws = ws_dash

    ws.sheet_view.showGridLines = False

    # --------------------------------------------------------
    # TÍTULO
    # --------------------------------------------------------

    ws.merge_cells(
        "A1:N2"
    )

    titulo = ws["A1"]

    titulo.value = (
        "RADAR DE TURISMO DE NATURALEZA Y AVENTURA DEL PERÚ"
    )

    titulo.font = Font(
        size=22,
        bold=True,
        color=BLANCO
    )

    titulo.fill = PatternFill(
        "solid",
        fgColor=VERDE_OSCURO
    )

    titulo.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    ws.merge_cells(
        "A3:N3"
    )

    ws["A3"] = (
        "Inversión pública, potencial territorial, "
        "ejecución y brechas de gestión | 2026"
    )

    ws["A3"].font = Font(
        size=11,
        italic=True,
        color=GRIS_OSCURO
    )

    ws["A3"].alignment = Alignment(
        horizontal="center"
    )

    # ========================================================
    # KPIs
    # ========================================================

    pim_total = ranking[
        "PIM_Radar"
    ].sum()

    territorios = len(
        ranking
    )

    lider = (
        ranking
        .sort_values(
            "IRNA_Estructural",
            ascending=False
        )
        .iloc[0]
    )

    consolidado = (
        ranking[
            ranking[
                "Estado_Territorial"
            ]
            == "FORTALEZA CONSOLIDADA"
        ]
        .sort_values(
            "IRNA_Estructural",
            ascending=False
        )
    )

    if len(consolidado) > 0:

        territorio_consolidado = (
            consolidado.iloc[0][
                "Departamento"
            ]
        )

    else:

        territorio_consolidado = "—"

    alertas_rojas = (
        ranking[
            "Nivel_Alerta"
        ]
        == "ROJA"
    ).sum()

    kpis = [
        (
            "A5:C8",
            "TERRITORIOS ANALIZADOS",
            territorios,
            AZUL
        ),
        (
            "D5:F8",
            "PIM NÚCLEO RADAR",
            f"S/ {pim_total/1_000_000:.1f} M",
            VERDE
        ),
        (
            "G5:I8",
            "LÍDER ESTRUCTURAL",
            lider["Departamento"],
            VERDE_OSCURO
        ),
        (
            "J5:L8",
            "FORTALEZA CONSOLIDADA",
            territorio_consolidado,
            AZUL
        ),
        (
            "M5:N8",
            "ALERTAS ROJAS",
            int(alertas_rojas),
            ROJO
        )
    ]

    for rango, etiqueta, valor, color in kpis:

        ws.merge_cells(
            rango
        )

        inicio = rango.split(
            ":"
        )[0]

        celda = ws[
            inicio
        ]

        celda.value = (
            f"{etiqueta}\n{valor}"
        )

        celda.font = Font(
            size=14,
            bold=True,
            color=BLANCO
        )

        celda.fill = PatternFill(
            "solid",
            fgColor=color
        )

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    # ========================================================
    # TOP 10
    # ========================================================

    ws["A10"] = (
        "TOP 10 — FORTALEZA ESTRUCTURAL"
    )

    ws["A10"].font = Font(
        size=14,
        bold=True,
        color=VERDE_OSCURO
    )

    ws["A11"] = "Departamento"
    ws["B11"] = "IRNA"
    ws["C11"] = "Ejecución"
    ws["D11"] = "Brecha"
    ws["E11"] = "Alerta"

    for cell in ws[
        "A11:E11"
    ][0]:

        cell.fill = PatternFill(
            "solid",
            fgColor=VERDE_OSCURO
        )

        cell.font = Font(
            bold=True,
            color=BLANCO
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    top10 = (
        ranking
        .sort_values(
            "IRNA_Estructural",
            ascending=False
        )
        .head(10)
    )

    fila = 12

    for _, row in top10.iterrows():

        ws.cell(
            fila,
            1,
            row["Departamento"]
        )

        ws.cell(
            fila,
            2,
            row["IRNA_Estructural"]
        )

        ws.cell(
            fila,
            3,
            row["IRNA_Ejecucion"]
        )

        ws.cell(
            fila,
            4,
            row["Brecha_IRNA"]
        )

        alerta = ws.cell(
            fila,
            5,
            row["Nivel_Alerta"]
        )

        alerta.fill = PatternFill(
            "solid",
            fgColor=color_alerta(
                row["Nivel_Alerta"]
            )
        )

        alerta.font = Font(
            bold=True,
            color=BLANCO
        )

        alerta.alignment = Alignment(
            horizontal="center"
        )

        fila += 1

    # --------------------------------------------------------
    # ESCALA DE COLOR IRNA
    # --------------------------------------------------------

    ws.conditional_formatting.add(
        "B12:B21",
        ColorScaleRule(
            start_type="min",
            start_color="FCE8E6",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFF4CC",
            end_type="max",
            end_color="D9EAD3"
        )
    )

    # ========================================================
    # GRÁFICO RANKING
    # ========================================================

    chart = BarChart()

    chart.type = "bar"

    chart.style = 10

    chart.title = (
        "Top 10 IRNA Estructural"
    )

    chart.y_axis.title = (
        "Departamento"
    )

    chart.x_axis.title = (
        "IRNA"
    )

    data = Reference(
        ws,
        min_col=2,
        min_row=11,
        max_row=21
    )

    cats = Reference(
        ws,
        min_col=1,
        min_row=12,
        max_row=21
    )

    chart.add_data(
        data,
        titles_from_data=True
    )

    chart.set_categories(
        cats
    )

    chart.height = 8
    chart.width = 13

    chart.legend = None

    chart.dataLabels = (
        DataLabelList()
    )

    chart.dataLabels.showVal = True

    ws.add_chart(
        chart,
        "G10"
    )

    # ========================================================
    # BRECHAS
    # ========================================================

    ws["A24"] = (
        "MAYORES BRECHAS ENTRE POTENCIAL Y EJECUCIÓN"
    )

    ws["A24"].font = Font(
        size=14,
        bold=True,
        color=ROJO
    )

    ws["A25"] = "Departamento"
    ws["B25"] = "IRNA Estructural"
    ws["C25"] = "Ejecución"
    ws["D25"] = "Brecha"

    for cell in ws[
        "A25:D25"
    ][0]:

        cell.fill = PatternFill(
            "solid",
            fgColor=ROJO
        )

        cell.font = Font(
            bold=True,
            color=BLANCO
        )

    top_brechas = (
        ranking
        .sort_values(
            "Brecha_IRNA",
            ascending=False
        )
        .head(8)
    )

    fila = 26

    for _, row in top_brechas.iterrows():

        ws.cell(
            fila,
            1,
            row["Departamento"]
        )

        ws.cell(
            fila,
            2,
            row["IRNA_Estructural"]
        )

        ws.cell(
            fila,
            3,
            row["IRNA_Ejecucion"]
        )

        ws.cell(
            fila,
            4,
            row["Brecha_IRNA"]
        )

        fila += 1

    # ========================================================
    # GRÁFICO BRECHAS
    # ========================================================

    chart_brecha = BarChart()

    chart_brecha.type = "bar"
    chart_brecha.style = 10

    chart_brecha.title = (
        "Principales brechas territoriales"
    )

    chart_brecha.x_axis.title = (
        "Brecha IRNA"
    )

    data = Reference(
        ws,
        min_col=4,
        min_row=25,
        max_row=33
    )

    cats = Reference(
        ws,
        min_col=1,
        min_row=26,
        max_row=33
    )

    chart_brecha.add_data(
        data,
        titles_from_data=True
    )

    chart_brecha.set_categories(
        cats
    )

    chart_brecha.height = 7
    chart_brecha.width = 13

    chart_brecha.legend = None

    ws.add_chart(
        chart_brecha,
        "G25"
    )

    # ========================================================
    # MATRIZ POTENCIAL VS EJECUCIÓN
    # ========================================================

    ws_scatter = wb[
        "Potencial_Ejecucion"
    ]

    scatter = ScatterChart()

    scatter.title = (
        "IRNA estructural vs ejecución"
    )

    scatter.x_axis.title = (
        "IRNA Estructural"
    )

    scatter.y_axis.title = (
        "IRNA Ejecución"
    )

    scatter.height = 10
    scatter.width = 16

    xvalues = Reference(
        ws_scatter,
        min_col=2,
        min_row=2,
        max_row=ws_scatter.max_row
    )

    yvalues = Reference(
        ws_scatter,
        min_col=3,
        min_row=2,
        max_row=ws_scatter.max_row
    )

    serie = Series(
        yvalues,
        xvalues,
        title="Territorios"
    )

    scatter.series.append(
        serie
    )

    ws.add_chart(
        scatter,
        "A36"
    )

    # ========================================================
    # NOTA METODOLÓGICA
    # ========================================================

    ws.merge_cells(
        "A56:N60"
    )

    ws["A56"] = (
        "LECTURA DEL RADAR\n"
        "IRNA Estructural mide fortaleza territorial, masa crítica de inversión, "
        "vocación, evidencia y diversidad. IRNA Ejecución mide el avance presupuestal. "
        "La Brecha IRNA identifica territorios donde el potencial estructural supera "
        "significativamente la capacidad actual de ejecución. Los cuadrantes se "
        "interpretan respecto de las medianas nacionales y no como umbrales absolutos."
    )

    ws["A56"].fill = PatternFill(
        "solid",
        fgColor=GRIS_CLARO
    )

    ws["A56"].font = Font(
        size=10,
        color=GRIS_OSCURO
    )

    ws["A56"].alignment = Alignment(
        vertical="top",
        wrap_text=True
    )

    # ========================================================
    # FORMATO GENERAL DASHBOARD
    # ========================================================

    for col, ancho in {
        "A": 24,
        "B": 14,
        "C": 14,
        "D": 14,
        "E": 14,
        "F": 3,
        "G": 14,
        "H": 14,
        "I": 14,
        "J": 14,
        "K": 14,
        "L": 14,
        "M": 14,
        "N": 14
    }.items():

        ws.column_dimensions[
            col
        ].width = ancho

    for row in range(
        1,
        61
    ):
        ws.row_dimensions[
            row
        ].height = 22

    ws.row_dimensions[
        1
    ].height = 32

    ws.freeze_panes = (
        "A10"
    )

    # ========================================================
    # FORMATEAR HOJAS DE DATOS
    # ========================================================

    for nombre in [
        "Ranking",
        "Potencial_Ejecucion",
        "Alertas",
        "Oportunidades",
        "Indicadores"
    ]:

        hoja = wb[
            nombre
        ]

        hoja.freeze_panes = (
            "A2"
        )

        hoja.auto_filter.ref = (
            hoja.dimensions
        )

        estilo_encabezado(
            hoja
        )

        ajustar_columnas(
            hoja
        )

        for row in hoja.iter_rows(
            min_row=2
        ):

            for cell in row:

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )

    # --------------------------------------------------------
    # FORMATOS NUMÉRICOS RANKING
    # --------------------------------------------------------

    hoja = wb[
        "Ranking"
    ]

    encabezados = {
        cell.value: cell.column
        for cell in hoja[1]
    }

    if "PIM_Radar" in encabezados:

        col = encabezados[
            "PIM_Radar"
        ]

        for fila in range(
            2,
            hoja.max_row + 1
        ):

            formato_moneda(
                hoja.cell(
                    fila,
                    col
                )
            )

    # ========================================================
    # GUARDAR
    # ========================================================

    wb.save(
        SALIDA
    )

    print("\n" + "=" * 100)
    print("DASHBOARD GENERADO")
    print("=" * 100)

    print(SALIDA)

    print(
        "\n✓ DASHBOARD EJECUTIVO DEL RADAR COMPLETADO"
    )

    print("FIN")


if __name__ == "__main__":
    main()