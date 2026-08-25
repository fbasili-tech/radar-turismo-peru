from pathlib import Path
import warnings

import numpy as np
import pandas as pd


# ============================================================
# RADAR TURISMO NATURALEZA Y AVENTURA - PERÚ
# 24 - DIAGNÓSTICO ESTADÍSTICO PARA IRNA-C
#
# OBJETIVO
# ------------------------------------------------------------
# Analizar la matriz maestra de Fase 2 antes de construir
# cualquier índice compuesto.
#
# Este programa:
# 1. Audita cobertura y calidad.
# 2. Describe las variables.
# 3. Detecta valores constantes y baja variabilidad.
# 4. Detecta valores extremos.
# 5. Calcula correlaciones.
# 6. Detecta pares altamente correlacionados.
# 7. Propone variables candidatas.
#
# IMPORTANTE:
# NO calcula todavía IRNA-C.
# NO asigna pesos.
# NO genera ranking.
# ============================================================


warnings.filterwarnings(
    "ignore",
    category=FutureWarning
)


# ============================================================
# ARCHIVOS
# ============================================================

ENTRADA = Path(
    "outputs/radar_fase2_matriz_maestra_2026.xlsx"
)

SALIDA = Path(
    "outputs/radar_fase2_diagnostico_irnac_2026.xlsx"
)

SALIDA.parent.mkdir(
    parents=True,
    exist_ok=True
)


# ============================================================
# CONFIGURACIÓN
# ============================================================

HOJA_ENTRADA = "02_Matriz_Analitica"

UMBRAL_CORRELACION_ALTA = 0.85
UMBRAL_CORRELACION_MEDIA = 0.70
UMBRAL_COBERTURA_MINIMA = 0.80

# Una variable con demasiados ceros puede ser poco adecuada
# para un índice continuo, aunque siga siendo informativa.
UMBRAL_CEROS_ADVERTENCIA = 0.60


# ============================================================
# DIMENSIONES CONCEPTUALES
# ============================================================

def detectar_dimension(variable):

    v = str(variable)

    if v.startswith("Base_"):
        return "IRNA_BASE"

    if v.startswith("Demanda_"):
        return "DEMANDA"

    if v.startswith("Hosp_"):
        return "HOSPEDAJE"

    if v.startswith("Aereo_"):
        return "CONECTIVIDAD"

    if v.startswith("Oferta_"):
        return "OFERTA_FORMAL"

    if v.startswith("Capital_"):
        return "CAPITAL_NATURAL"

    if v.startswith("Deriv_"):
        return "VARIABLE_DERIVADA"

    return "OTRA"


# ============================================================
# VARIABLES QUE NO DEBEN ENTRAR DIRECTAMENTE AL IRNA-C
# ============================================================

def es_variable_base_excluida(variable):

    """
    El IRNA-C debe construirse inicialmente con variables
    observadas de Fase 2.

    Los IRNA previos se conservan para comparación posterior,
    no para introducir circularidad dentro del nuevo índice.
    """

    return str(variable).startswith(
        "Base_"
    )


# ============================================================
# LECTURA
# ============================================================

def cargar_datos():

    if not ENTRADA.exists():

        raise FileNotFoundError(
            f"No existe:\n{ENTRADA}"
        )

    xls = pd.ExcelFile(
        ENTRADA
    )

    print(
        "Hojas disponibles:"
    )

    print(
        xls.sheet_names
    )

    if HOJA_ENTRADA not in xls.sheet_names:

        raise ValueError(
            f"No existe la hoja "
            f"{HOJA_ENTRADA}"
        )

    df = pd.read_excel(
        ENTRADA,
        sheet_name=HOJA_ENTRADA
    )

    if "Departamento" not in df.columns:

        raise ValueError(
            "La matriz no contiene Departamento."
        )

    print(
        f"\nTerritorios cargados : {len(df)}"
    )

    print(
        f"Variables cargadas   : "
        f"{len(df.columns) - 1}"
    )

    return df


# ============================================================
# IDENTIFICAR VARIABLES NUMÉRICAS
# ============================================================

def obtener_variables_numericas(df):

    variables = []

    for columna in df.columns:

        if columna == "Departamento":
            continue

        serie = pd.to_numeric(
            df[columna],
            errors="coerce"
        )

        if serie.notna().sum() > 0:

            variables.append(
                columna
            )

    return variables


# ============================================================
# AUDITORÍA VARIABLE POR VARIABLE
# ============================================================

def auditar_variables(
    df,
    variables
):

    filas = []

    n = len(df)

    for variable in variables:

        s = pd.to_numeric(
            df[variable],
            errors="coerce"
        )

        validos = int(
            s.notna().sum()
        )

        faltantes = int(
            s.isna().sum()
        )

        ceros = int(
            (s.fillna(np.nan) == 0).sum()
        )

        cobertura = (
            validos / n
            if n
            else np.nan
        )

        proporcion_ceros = (
            ceros / validos
            if validos
            else np.nan
        )

        unicos = int(
            s.dropna().nunique()
        )

        minimo = (
            s.min()
            if validos
            else np.nan
        )

        maximo = (
            s.max()
            if validos
            else np.nan
        )

        media = (
            s.mean()
            if validos
            else np.nan
        )

        mediana = (
            s.median()
            if validos
            else np.nan
        )

        desviacion = (
            s.std()
            if validos > 1
            else np.nan
        )

        if (
            pd.notna(media)
            and media != 0
            and pd.notna(desviacion)
        ):

            coef_variacion = (
                desviacion
                / abs(media)
            )

        else:

            coef_variacion = np.nan

        constante = int(
            unicos <= 1
        )

        dimension = detectar_dimension(
            variable
        )

        excluir_base = (
            es_variable_base_excluida(
                variable
            )
        )

        observaciones = []

        if cobertura < UMBRAL_COBERTURA_MINIMA:

            observaciones.append(
                "Cobertura baja"
            )

        if constante:

            observaciones.append(
                "Variable constante"
            )

        if (
            pd.notna(
                proporcion_ceros
            )
            and
            proporcion_ceros
            >= UMBRAL_CEROS_ADVERTENCIA
        ):

            observaciones.append(
                "Alta proporción de ceros"
            )

        if excluir_base:

            observaciones.append(
                "Variable base: usar para contraste, no dentro del IRNA-C"
            )

        filas.append(
            {
                "Variable":
                    variable,

                "Dimension":
                    dimension,

                "Territorios_Validos":
                    validos,

                "Faltantes":
                    faltantes,

                "Cobertura":
                    cobertura,

                "Ceros":
                    ceros,

                "Proporcion_Ceros":
                    proporcion_ceros,

                "Valores_Unicos":
                    unicos,

                "Minimo":
                    minimo,

                "Maximo":
                    maximo,

                "Media":
                    media,

                "Mediana":
                    mediana,

                "Desviacion_Estandar":
                    desviacion,

                "Coeficiente_Variacion":
                    coef_variacion,

                "Es_Constante":
                    constante,

                "Excluir_Por_Ser_Base":
                    int(
                        excluir_base
                    ),

                "Observacion":
                    " | ".join(
                        observaciones
                    )
            }
        )

    return pd.DataFrame(
        filas
    )


# ============================================================
# OUTLIERS - MÉTODO IQR
# ============================================================

def detectar_outliers(
    df,
    variables
):

    detalle = []
    resumen = []

    for variable in variables:

        s = pd.to_numeric(
            df[variable],
            errors="coerce"
        )

        valores = s.dropna()

        if len(valores) < 4:

            continue

        q1 = valores.quantile(
            0.25
        )

        q3 = valores.quantile(
            0.75
        )

        iqr = q3 - q1

        if pd.isna(iqr):

            continue

        limite_inferior = (
            q1 - 1.5 * iqr
        )

        limite_superior = (
            q3 + 1.5 * iqr
        )

        mascara = (
            (s < limite_inferior)
            |
            (s > limite_superior)
        )

        cantidad = int(
            mascara.sum()
        )

        resumen.append(
            {
                "Variable":
                    variable,

                "Dimension":
                    detectar_dimension(
                        variable
                    ),

                "Q1":
                    q1,

                "Q3":
                    q3,

                "IQR":
                    iqr,

                "Limite_Inferior":
                    limite_inferior,

                "Limite_Superior":
                    limite_superior,

                "Outliers":
                    cantidad,

                "Proporcion_Outliers":
                    (
                        cantidad
                        / s.notna().sum()
                        if s.notna().sum()
                        else np.nan
                    )
            }
        )

        if cantidad > 0:

            for indice in df.index[
                mascara.fillna(False)
            ]:

                detalle.append(
                    {
                        "Departamento":
                            df.loc[
                                indice,
                                "Departamento"
                            ],

                        "Variable":
                            variable,

                        "Dimension":
                            detectar_dimension(
                                variable
                            ),

                        "Valor":
                            s.loc[
                                indice
                            ],

                        "Limite_Inferior":
                            limite_inferior,

                        "Limite_Superior":
                            limite_superior
                    }
                )

    return (
        pd.DataFrame(
            resumen
        ),
        pd.DataFrame(
            detalle
        )
    )


# ============================================================
# CORRELACIONES
# ============================================================

def calcular_correlaciones(
    df,
    variables
):

    numerico = pd.DataFrame(
        index=df.index
    )

    for variable in variables:

        numerico[
            variable
        ] = pd.to_numeric(
            df[
                variable
            ],
            errors="coerce"
        )

    correlacion = numerico.corr(
        method="spearman",
        min_periods=10
    )

    return correlacion


# ============================================================
# PARES CORRELACIONADOS
# ============================================================

def construir_pares_correlacion(
    correlacion
):

    filas = []

    columnas = list(
        correlacion.columns
    )

    for i in range(
        len(columnas)
    ):

        for j in range(
            i + 1,
            len(columnas)
        ):

            a = columnas[i]
            b = columnas[j]

            r = correlacion.loc[
                a,
                b
            ]

            if pd.isna(r):
                continue

            abs_r = abs(r)

            if (
                abs_r
                >= UMBRAL_CORRELACION_ALTA
            ):

                nivel = "ALTA"

            elif (
                abs_r
                >= UMBRAL_CORRELACION_MEDIA
            ):

                nivel = "MEDIA"

            else:

                nivel = "BAJA"

            filas.append(
                {
                    "Variable_1":
                        a,

                    "Dimension_1":
                        detectar_dimension(
                            a
                        ),

                    "Variable_2":
                        b,

                    "Dimension_2":
                        detectar_dimension(
                            b
                        ),

                    "Correlacion_Spearman":
                        r,

                    "Correlacion_Absoluta":
                        abs_r,

                    "Nivel":
                        nivel,

                    "Misma_Dimension":
                        int(
                            detectar_dimension(a)
                            ==
                            detectar_dimension(b)
                        )
                }
            )

    resultado = pd.DataFrame(
        filas
    )

    if not resultado.empty:

        resultado = resultado.sort_values(
            "Correlacion_Absoluta",
            ascending=False
        )

    return resultado


# ============================================================
# PROPUESTA PRELIMINAR DE VARIABLES
# ============================================================

def construir_candidatas(
    auditoria,
    pares
):

    resultado = auditoria.copy()

    resultado[
        "Correlaciones_Altas"
    ] = 0

    resultado[
        "Correlaciones_Altas_Misma_Dimension"
    ] = 0

    if not pares.empty:

        altas = pares[
            pares[
                "Nivel"
            ] == "ALTA"
        ]

        for variable in resultado[
            "Variable"
        ]:

            relacionadas = altas[
                (
                    altas[
                        "Variable_1"
                    ]
                    == variable
                )
                |
                (
                    altas[
                        "Variable_2"
                    ]
                    == variable
                )
            ]

            resultado.loc[
                resultado[
                    "Variable"
                ]
                == variable,
                "Correlaciones_Altas"
            ] = len(
                relacionadas
            )

            misma = relacionadas[
                relacionadas[
                    "Misma_Dimension"
                ]
                == 1
            ]

            resultado.loc[
                resultado[
                    "Variable"
                ]
                == variable,
                "Correlaciones_Altas_Misma_Dimension"
            ] = len(
                misma
            )

    decisiones = []
    razones = []

    for _, fila in resultado.iterrows():

        variable = fila[
            "Variable"
        ]

        cobertura = fila[
            "Cobertura"
        ]

        constante = fila[
            "Es_Constante"
        ]

        excluir_base = fila[
            "Excluir_Por_Ser_Base"
        ]

        proporcion_ceros = fila[
            "Proporcion_Ceros"
        ]

        correlaciones = fila[
            "Correlaciones_Altas_Misma_Dimension"
        ]

        if excluir_base:

            decision = "CONTRASTE"

            razon = (
                "IRNA previo; conservar para "
                "comparación externa."
            )

        elif constante == 1:

            decision = "EXCLUIR"

            razon = (
                "No discrimina entre territorios."
            )

        elif cobertura < UMBRAL_COBERTURA_MINIMA:

            decision = "EVALUAR"

            razon = (
                "Cobertura territorial insuficiente."
            )

        elif (
            pd.notna(
                proporcion_ceros
            )
            and
            proporcion_ceros
            >= UMBRAL_CEROS_ADVERTENCIA
        ):

            decision = "EVALUAR"

            razon = (
                "Alta concentración de valores cero."
            )

        elif correlaciones >= 2:

            decision = "EVALUAR"

            razon = (
                "Alta redundancia con variables "
                "de la misma dimensión."
            )

        else:

            decision = "CANDIDATA"

            razon = (
                "Cobertura y variabilidad adecuadas."
            )

        decisiones.append(
            decision
        )

        razones.append(
            razon
        )

    resultado[
        "Decision_Preliminar"
    ] = decisiones

    resultado[
        "Razon_Decision"
    ] = razones

    return resultado


# ============================================================
# RESUMEN POR DIMENSIÓN
# ============================================================

def construir_resumen_dimensiones(
    candidatas
):

    return (
        candidatas
        .groupby(
            "Dimension",
            as_index=False
        )
        .agg(
            Variables=(
                "Variable",
                "count"
            ),

            Candidatas=(
                "Decision_Preliminar",
                lambda x:
                (
                    x
                    == "CANDIDATA"
                ).sum()
            ),

            Evaluar=(
                "Decision_Preliminar",
                lambda x:
                (
                    x
                    == "EVALUAR"
                ).sum()
            ),

            Excluir=(
                "Decision_Preliminar",
                lambda x:
                (
                    x
                    == "EXCLUIR"
                ).sum()
            ),

            Contraste=(
                "Decision_Preliminar",
                lambda x:
                (
                    x
                    == "CONTRASTE"
                ).sum()
            ),

            Cobertura_Media=(
                "Cobertura",
                "mean"
            )
        )
    )


# ============================================================
# TOP TERRITORIOS POR VARIABLE
# ============================================================

def construir_top_variables(
    df,
    candidatas
):

    filas = []

    variables = candidatas.loc[
        candidatas[
            "Decision_Preliminar"
        ].isin(
            [
                "CANDIDATA",
                "EVALUAR"
            ]
        ),
        "Variable"
    ].tolist()

    for variable in variables:

        s = pd.to_numeric(
            df[
                variable
            ],
            errors="coerce"
        )

        temporal = pd.DataFrame(
            {
                "Departamento":
                    df[
                        "Departamento"
                    ],

                "Valor":
                    s
            }
        )

        temporal = (
            temporal
            .dropna()
            .sort_values(
                "Valor",
                ascending=False
            )
            .head(5)
        )

        for posicion, (
            _,
            fila
        ) in enumerate(
            temporal.iterrows(),
            start=1
        ):

            filas.append(
                {
                    "Variable":
                        variable,

                    "Dimension":
                        detectar_dimension(
                            variable
                        ),

                    "Posicion":
                        posicion,

                    "Departamento":
                        fila[
                            "Departamento"
                        ],

                    "Valor":
                        fila[
                            "Valor"
                        ]
                }
            )

    return pd.DataFrame(
        filas
    )


# ============================================================
# METODOLOGÍA
# ============================================================

def construir_metodologia():

    datos = [
        (
            "Propósito",
            "Diagnosticar estadísticamente las variables antes de construir el IRNA-C."
        ),
        (
            "Unidad de análisis",
            "25 territorios del Radar."
        ),
        (
            "Correlación",
            "Se utiliza Spearman por el tamaño de muestra y por no asumir normalidad."
        ),
        (
            "Correlación alta",
            "Valor absoluto igual o superior a 0.85."
        ),
        (
            "Correlación media",
            "Valor absoluto entre 0.70 y 0.85."
        ),
        (
            "Outliers",
            "Se utiliza la regla IQR de 1.5 veces el rango intercuartílico."
        ),
        (
            "IRNA previo",
            "Las variables Base_IRNA se conservan para contraste y no ingresan automáticamente al IRNA-C."
        ),
        (
            "Ceros",
            "Un valor cero no se interpreta automáticamente como dato faltante."
        ),
        (
            "Variables derivadas",
            "Se analizan, pero deben justificarse conceptualmente antes de ingresar al índice."
        ),
        (
            "Pesos",
            "Esta etapa no asigna pesos."
        ),
        (
            "Ranking",
            "Esta etapa no genera ranking territorial."
        ),
        (
            "Siguiente etapa",
            "Seleccionar indicadores definitivos, sentido de cada indicador y método de normalización."
        )
    ]

    return pd.DataFrame(
        datos,
        columns=[
            "Tema",
            "Criterio"
        ]
    )


# ============================================================
# FORMATO EXCEL
# ============================================================

def formatear_excel(ruta):

    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(
        ruta
    )

    for ws in wb.worksheets:

        ws.freeze_panes = "A2"

        if (
            ws.max_row > 1
            and
            ws.max_column > 1
        ):

            ws.auto_filter.ref = (
                ws.dimensions
            )

        for celda in ws[1]:

            celda.font = Font(
                bold=True
            )

            celda.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        for columna in range(
            1,
            ws.max_column + 1
        ):

            letra = get_column_letter(
                columna
            )

            maximo = 0

            for fila in range(
                1,
                min(
                    ws.max_row,
                    150
                ) + 1
            ):

                valor = ws.cell(
                    fila,
                    columna
                ).value

                if valor is not None:

                    maximo = max(
                        maximo,
                        len(
                            str(valor)
                        )
                    )

            ws.column_dimensions[
                letra
            ].width = min(
                max(
                    maximo + 2,
                    12
                ),
                42
            )

    wb.save(
        ruta
    )


# ============================================================
# MAIN
# ============================================================

def main():

    print(
        "=" * 100
    )

    print(
        "RADAR TURISMO NATURALEZA "
        "Y AVENTURA - PERÚ"
    )

    print(
        "24 - DIAGNÓSTICO ESTADÍSTICO IRNA-C"
    )

    print(
        "=" * 100
    )

    df = cargar_datos()

    variables = obtener_variables_numericas(
        df
    )

    print(
        f"Variables numéricas : "
        f"{len(variables)}"
    )

    # --------------------------------------------------------
    # AUDITORÍA
    # --------------------------------------------------------

    auditoria = auditar_variables(
        df,
        variables
    )

    # --------------------------------------------------------
    # OUTLIERS
    # --------------------------------------------------------

    resumen_outliers, detalle_outliers = (
        detectar_outliers(
            df,
            variables
        )
    )

    # --------------------------------------------------------
    # CORRELACIONES
    # --------------------------------------------------------

    correlacion = calcular_correlaciones(
        df,
        variables
    )

    pares = construir_pares_correlacion(
        correlacion
    )

    # --------------------------------------------------------
    # CANDIDATAS
    # --------------------------------------------------------

    candidatas = construir_candidatas(
        auditoria,
        pares
    )

    dimensiones = construir_resumen_dimensiones(
        candidatas
    )

    top_variables = construir_top_variables(
        df,
        candidatas
    )

    metodologia = construir_metodologia()

    # --------------------------------------------------------
    # RESUMEN GENERAL
    # --------------------------------------------------------

    n_candidatas = int(
        (
            candidatas[
                "Decision_Preliminar"
            ]
            == "CANDIDATA"
        ).sum()
    )

    n_evaluar = int(
        (
            candidatas[
                "Decision_Preliminar"
            ]
            == "EVALUAR"
        ).sum()
    )

    n_excluir = int(
        (
            candidatas[
                "Decision_Preliminar"
            ]
            == "EXCLUIR"
        ).sum()
    )

    n_contraste = int(
        (
            candidatas[
                "Decision_Preliminar"
            ]
            == "CONTRASTE"
        ).sum()
    )

    if pares.empty:

        correlaciones_altas = 0
        correlaciones_medias = 0

    else:

        correlaciones_altas = int(
            (
                pares[
                    "Nivel"
                ]
                == "ALTA"
            ).sum()
        )

        correlaciones_medias = int(
            (
                pares[
                    "Nivel"
                ]
                == "MEDIA"
            ).sum()
        )

    total_outliers = (
        len(
            detalle_outliers
        )
    )

    resumen_general = pd.DataFrame(
        {
            "Indicador": [
                "Territorios",
                "Variables numéricas",
                "Variables candidatas",
                "Variables a evaluar",
                "Variables excluidas",
                "Variables de contraste",
                "Pares con correlación alta",
                "Pares con correlación media",
                "Observaciones outlier"
            ],

            "Valor": [
                len(
                    df
                ),
                len(
                    variables
                ),
                n_candidatas,
                n_evaluar,
                n_excluir,
                n_contraste,
                correlaciones_altas,
                correlaciones_medias,
                total_outliers
            ]
        }
    )

    # --------------------------------------------------------
    # EXPORTAR
    # --------------------------------------------------------

    with pd.ExcelWriter(
        SALIDA,
        engine="openpyxl"
    ) as writer:

        resumen_general.to_excel(
            writer,
            sheet_name="01_Resumen",
            index=False
        )

        candidatas.to_excel(
            writer,
            sheet_name="02_Diagnostico_Variables",
            index=False
        )

        dimensiones.to_excel(
            writer,
            sheet_name="03_Dimensiones",
            index=False
        )

        correlacion.to_excel(
            writer,
            sheet_name="04_Matriz_Correlacion"
        )

        pares.to_excel(
            writer,
            sheet_name="05_Pares_Correlacion",
            index=False
        )

        resumen_outliers.to_excel(
            writer,
            sheet_name="06_Resumen_Outliers",
            index=False
        )

        detalle_outliers.to_excel(
            writer,
            sheet_name="07_Detalle_Outliers",
            index=False
        )

        top_variables.to_excel(
            writer,
            sheet_name="08_Top_Variables",
            index=False
        )

        metodologia.to_excel(
            writer,
            sheet_name="09_Metodologia",
            index=False
        )

    formatear_excel(
        SALIDA
    )

    # --------------------------------------------------------
    # SALIDA CONSOLA
    # --------------------------------------------------------

    print(
        "\n" + "=" * 100
    )

    print(
        "RESUMEN DIAGNÓSTICO"
    )

    print(
        "=" * 100
    )

    print(
        resumen_general.to_string(
            index=False
        )
    )

    print(
        "\n" + "=" * 100
    )

    print(
        "RESUMEN POR DIMENSIÓN"
    )

    print(
        "=" * 100
    )

    print(
        dimensiones.to_string(
            index=False,
            formatters={
                "Cobertura_Media":
                    lambda x:
                    f"{x:.1%}"
            }
        )
    )

    print(
        "\n" + "=" * 100
    )

    print(
        "PARES CON CORRELACIÓN ALTA"
    )

    print(
        "=" * 100
    )

    if pares.empty:

        print(
            "No se detectaron pares."
        )

    else:

        altas = pares[
            pares[
                "Nivel"
            ]
            == "ALTA"
        ]

        if altas.empty:

            print(
                "No se detectaron "
                "correlaciones altas."
            )

        else:

            print(
                altas[
                    [
                        "Variable_1",
                        "Variable_2",
                        "Correlacion_Spearman",
                        "Misma_Dimension"
                    ]
                ]
                .to_string(
                    index=False
                )
            )

    print(
        "\n" + "=" * 100
    )

    print(
        "DECISIÓN PRELIMINAR"
    )

    print(
        "=" * 100
    )

    print(
        candidatas[
            [
                "Variable",
                "Dimension",
                "Cobertura",
                "Proporcion_Ceros",
                "Correlaciones_Altas_Misma_Dimension",
                "Decision_Preliminar",
                "Razon_Decision"
            ]
        ]
        .to_string(
            index=False,
            formatters={
                "Cobertura":
                    lambda x:
                    f"{x:.1%}",

                "Proporcion_Ceros":
                    lambda x:
                    (
                        f"{x:.1%}"
                        if pd.notna(x)
                        else ""
                    )
            }
        )
    )

    print(
        "\nARCHIVO GENERADO:"
    )

    print(
        SALIDA
    )

    print(
        "\n✓ DIAGNÓSTICO ESTADÍSTICO "
        "IRNA-C COMPLETADO"
    )

    print(
        "✓ NO SE HAN ASIGNADO PESOS"
    )

    print(
        "✓ NO SE HA GENERADO RANKING"
    )

    print(
        "FIN"
    )


if __name__ == "__main__":
    main()