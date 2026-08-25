from pathlib import Path
import pandas as pd
import numpy as np


# ============================================================
# RADAR TURISMO NATURALEZA Y AVENTURA - PERÚ
# FASE 2
#
# 23 - INTEGRADOR FASE 2
#
# OBJETIVO:
# Integrar en una sola matriz territorial:
#
# 18 Demanda turística
# 19 Hospedaje y permanencia
# 20 Conectividad aérea
# 21 Oferta turística formal
# 22 Capital natural / ANP
# + IRNA estructural previo
#
# Esta etapa NO calcula todavía el IRNA-C.
# ============================================================


# ============================================================
# ARCHIVOS DE ENTRADA
# ============================================================

ARCHIVO_IRNA = Path(
    "outputs/radar_turismo_irna_calibrado_2026.xlsx"
)

ARCHIVO_DEMANDA = Path(
    "outputs/radar_fase2_demanda_turistica_2026.xlsx"
)

ARCHIVO_HOSPEDAJE = Path(
    "outputs/radar_fase2_hospedaje_permanencia_2026.xlsx"
)

ARCHIVO_AEREO = Path(
    "outputs/radar_fase2_conectividad_aerea_2026.xlsx"
)

ARCHIVO_OFERTA = Path(
    "outputs/radar_fase2_oferta_turistica_formal_2026.xlsx"
)

ARCHIVO_CAPITAL = Path(
    "outputs/radar_fase2_capital_natural_anp_2026.xlsx"
)


# ============================================================
# SALIDA
# ============================================================

SALIDA = Path(
    "outputs/radar_fase2_matriz_maestra_2026.xlsx"
)

SALIDA.parent.mkdir(
    parents=True,
    exist_ok=True
)


# ============================================================
# CATÁLOGO TERRITORIAL
# ============================================================

DEPARTAMENTOS_PERU = [
    "AMAZONAS",
    "ANCASH",
    "APURIMAC",
    "AREQUIPA",
    "AYACUCHO",
    "CAJAMARCA",
    "CUSCO",
    "HUANCAVELICA",
    "HUANUCO",
    "ICA",
    "JUNIN",
    "LA LIBERTAD",
    "LAMBAYEQUE",
    "LIMA",
    "LORETO",
    "MADRE DE DIOS",
    "MOQUEGUA",
    "PASCO",
    "PIURA",
    "PROVINCIA CONSTITUCIONAL DEL CALLAO",
    "PUNO",
    "SAN MARTIN",
    "TACNA",
    "TUMBES",
    "UCAYALI"
]


# ============================================================
# UTILIDADES
# ============================================================

def validar_archivo(ruta):

    if not ruta.exists():

        raise FileNotFoundError(
            f"No existe el archivo requerido:\n{ruta}"
        )


def normalizar_departamento(serie):

    resultado = (
        serie
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    resultado = (
        resultado
        .str.replace(
            "Á",
            "A",
            regex=False
        )
        .str.replace(
            "É",
            "E",
            regex=False
        )
        .str.replace(
            "Í",
            "I",
            regex=False
        )
        .str.replace(
            "Ó",
            "O",
            regex=False
        )
        .str.replace(
            "Ú",
            "U",
            regex=False
        )
    )

    resultado = resultado.replace(
        {
            "CALLAO":
                "PROVINCIA CONSTITUCIONAL DEL CALLAO",

            "PROVINCIA CONSTITUCIONAL CALLAO":
                "PROVINCIA CONSTITUCIONAL DEL CALLAO"
        }
    )

    return resultado


def seleccionar_columnas_existentes(
    df,
    columnas
):

    return [
        c
        for c in columnas
        if c in df.columns
    ]


def convertir_numericas(
    df,
    excluir=None
):

    excluir = excluir or []

    for columna in df.columns:

        if columna in excluir:
            continue

        df[columna] = pd.to_numeric(
            df[columna],
            errors="ignore"
        )

    return df


# ============================================================
# CARGAR IRNA
# ============================================================

def cargar_irna():

    validar_archivo(
        ARCHIVO_IRNA
    )

    print(
        "\nCargando IRNA estructural..."
    )

    xls = pd.ExcelFile(
        ARCHIVO_IRNA
    )

    print(
        "Hojas disponibles:"
    )

    print(
        xls.sheet_names
    )

    # Intentar encontrar hoja adecuada
    candidatos = [
        "IRNA_Calibrado",
        "Ranking_IRNA",
        "IRNA_Estructural",
        "Ranking",
        xls.sheet_names[0]
    ]

    hoja = None

    for candidato in candidatos:

        if candidato in xls.sheet_names:

            hoja = candidato
            break

    if hoja is None:

        hoja = xls.sheet_names[0]

    df = pd.read_excel(
        ARCHIVO_IRNA,
        sheet_name=hoja
    )

    print(
        f"Hoja utilizada: {hoja}"
    )

    if "Departamento" not in df.columns:

        raise ValueError(
            "No se encontró la columna Departamento "
            "en el archivo IRNA."
        )

    df[
        "Departamento"
    ] = normalizar_departamento(
        df[
            "Departamento"
        ]
    )

    columnas = seleccionar_columnas_existentes(
        df,
        [
            "Departamento",
            "IRNA_Estructural",
            "Categoria_Estructural",
            "IRNA_Ejecucion",
            "Categoria_Ejecucion",
            "Brecha_IRNA",
            "PIM_Radar",
            "Registros_Radar",
            "Proyectos_Alta_Vocacion",
            "Proyectos_Con_Evidencia"
        ]
    )

    df = df[
        columnas
    ].copy()

    prefijos = {
        c:
            (
                c
                if c == "Departamento"
                else "Base_" + c
            )
        for c in df.columns
    }

    df = df.rename(
        columns=prefijos
    )

    return df


# ============================================================
# CARGAR DEMANDA
# ============================================================

def cargar_demanda():

    validar_archivo(
        ARCHIVO_DEMANDA
    )

    print(
        "\nCargando demanda turística..."
    )

    df = pd.read_excel(
        ARCHIVO_DEMANDA,
        sheet_name="Demanda_Territorial"
    )

    df[
        "Departamento"
    ] = normalizar_departamento(
        df[
            "Departamento"
        ]
    )

    columnas = seleccionar_columnas_existentes(
        df,
        [
            "Departamento",
            "Visitantes_Sitios",
            "Visitantes_Nacionales",
            "Visitantes_Extranjeros",
            "Visitantes_Desagregados",
            "Diferencia_Total_Desagregacion",
            "Cobertura_Desagregacion_Porcentaje",
            "Participacion_Extranjera_Porcentaje",
            "Participacion_Nacional_Porcentaje",
            "Sitios_Con_Registro",
            "Anio_Demanda",
            "Tiene_Dato_Visitacion"
        ]
    )

    df = df[
        columnas
    ].copy()

    df = df.rename(
        columns={
            c:
                (
                    c
                    if c == "Departamento"
                    else "Demanda_" + c
                )
            for c in df.columns
        }
    )

    return df


# ============================================================
# CARGAR HOSPEDAJE
# ============================================================

def cargar_hospedaje():

    validar_archivo(
        ARCHIVO_HOSPEDAJE
    )

    print(
        "\nCargando hospedaje y permanencia..."
    )

    df = pd.read_excel(
        ARCHIVO_HOSPEDAJE,
        sheet_name="Hospedaje_Territorial"
    )

    df[
        "Departamento"
    ] = normalizar_departamento(
        df[
            "Departamento"
        ]
    )

    columnas = seleccionar_columnas_existentes(
        df,
        [
            "Departamento",
            "Meses_Con_Dato",
            "Arribos",
            "Arribos_Nacionales",
            "Arribos_Extranjeros",
            "Pernoctaciones",
            "Pernoctaciones_Nacionales",
            "Pernoctaciones_Extranjeros",
            "Permanencia_Promedio",
            "Permanencia_Nacionales",
            "Permanencia_Extranjeros",
            "TNOH_Promedio",
            "TNOC_Promedio",
            "Establecimientos_Promedio",
            "Habitaciones_Promedio",
            "Plazas_Cama_Promedio",
            "Empleo_Promedio",
            "Participacion_Arribos_Extranjeros",
            "Participacion_Pernoctaciones_Extranjeros",
            "Tiene_Dato_Hospedaje"
        ]
    )

    df = df[
        columnas
    ].copy()

    df = df.rename(
        columns={
            c:
                (
                    c
                    if c == "Departamento"
                    else "Hosp_" + c
                )
            for c in df.columns
        }
    )

    return df


# ============================================================
# CARGAR CONECTIVIDAD AÉREA
# ============================================================

def cargar_aereo():

    validar_archivo(
        ARCHIVO_AEREO
    )

    print(
        "\nCargando conectividad aérea..."
    )

    df = pd.read_excel(
        ARCHIVO_AEREO,
        sheet_name="Conectividad_Territorial"
    )

    df[
        "Departamento"
    ] = normalizar_departamento(
        df[
            "Departamento"
        ]
    )

    columnas = seleccionar_columnas_existentes(
        df,
        [
            "Departamento",
            "Pasajeros_Total",
            "Pasajeros_Domesticos",
            "Pasajeros_Internacionales",
            "Participacion_Internacional_Porcentaje",
            "Llegadas",
            "Salidas",
            "Aeropuertos_Con_Registro",
            "Aeropuertos_Internacionales",
            "Tiene_Conectividad_Aerea",
            "Tiene_Conectividad_Internacional",
            "Control_Llegadas_Salidas",
            "Control_Domestico_Internacional"
        ]
    )

    df = df[
        columnas
    ].copy()

    df = df.rename(
        columns={
            c:
                (
                    c
                    if c == "Departamento"
                    else "Aereo_" + c
                )
            for c in df.columns
        }
    )

    return df


# ============================================================
# CARGAR OFERTA FORMAL
# ============================================================

def cargar_oferta():

    validar_archivo(
        ARCHIVO_OFERTA
    )

    print(
        "\nCargando oferta turística formal..."
    )

    df = pd.read_excel(
        ARCHIVO_OFERTA,
        sheet_name="Oferta_Territorial"
    )

    df[
        "Departamento"
    ] = normalizar_departamento(
        df[
            "Departamento"
        ]
    )

    columnas = seleccionar_columnas_existentes(
        df,
        [
            "Departamento",
            "Ranking_Oferta",
            "Hospedajes_Formales",
            "Agencias_Formales",
            "Restaurantes_Calificados",
            "Prestadores_Formales_Total",
            "Tiene_Oferta_Formal"
        ]
    )

    df = df[
        columnas
    ].copy()

    df = df.rename(
        columns={
            c:
                (
                    c
                    if c == "Departamento"
                    else "Oferta_" + c
                )
            for c in df.columns
        }
    )

    return df


# ============================================================
# CARGAR CAPITAL NATURAL
# ============================================================

def cargar_capital():

    validar_archivo(
        ARCHIVO_CAPITAL
    )

    print(
        "\nCargando capital natural..."
    )

    df = pd.read_excel(
        ARCHIVO_CAPITAL,
        sheet_name="02_Capital_Territorial"
    )

    df[
        "Departamento"
    ] = normalizar_departamento(
        df[
            "Departamento"
        ]
    )

    columnas = seleccionar_columnas_existentes(
        df,
        [
            "Departamento",
            "ANP_Total",
            "ANP_Nacional",
            "ACR_Total",
            "ACP_Total",
            "Zonas_Reservadas",
            "ANP_Sin_Clasificar",
            "ANP_Multidepartamentales",
            "Diversidad_Categorias",
            "Superficie_Atribuida_Referencial_Ha",
            "Tiene_Capital_Natural"
        ]
    )

    df = df[
        columnas
    ].copy()

    df = df.rename(
        columns={
            c:
                (
                    c
                    if c == "Departamento"
                    else "Capital_" + c
                )
            for c in df.columns
        }
    )

    return df


# ============================================================
# INTEGRAR
# ============================================================

def construir_matriz_maestra():

    matriz = pd.DataFrame(
        {
            "Departamento":
                DEPARTAMENTOS_PERU
        }
    )

    capas = [
        cargar_irna(),
        cargar_demanda(),
        cargar_hospedaje(),
        cargar_aereo(),
        cargar_oferta(),
        cargar_capital()
    ]

    for capa in capas:

        matriz = matriz.merge(
            capa,
            on="Departamento",
            how="left",
            validate="one_to_one"
        )

    return matriz


# ============================================================
# CONTROL DE COBERTURA
# ============================================================

def construir_control_cobertura(
    matriz
):

    capas = {
        "IRNA":
            "Base_IRNA_Estructural",

        "Demanda":
            "Demanda_Visitantes_Sitios",

        "Hospedaje":
            "Hosp_Arribos",

        "Conectividad_Aerea":
            "Aereo_Pasajeros_Total",

        "Oferta_Formal":
            "Oferta_Prestadores_Formales_Total",

        "Capital_Natural":
            "Capital_ANP_Total"
    }

    filas = []

    for capa, columna in capas.items():

        if columna in matriz.columns:

            con_dato = int(
                matriz[
                    columna
                ]
                .notna()
                .sum()
            )

            ceros = int(
                (
                    pd.to_numeric(
                        matriz[
                            columna
                        ],
                        errors="coerce"
                    )
                    .fillna(0)
                    == 0
                ).sum()
            )

        else:

            con_dato = 0
            ceros = 25

        filas.append(
            {
                "Capa":
                    capa,

                "Territorios_Catalogo":
                    len(
                        matriz
                    ),

                "Territorios_Con_Registro":
                    con_dato,

                "Territorios_Sin_Registro":
                    len(
                        matriz
                    )
                    - con_dato,

                "Territorios_Valor_Cero":
                    ceros,

                "Cobertura_Porcentaje":
                    (
                        con_dato
                        / len(
                            matriz
                        )
                        * 100
                    )
            }
        )

    return pd.DataFrame(
        filas
    )


# ============================================================
# VARIABLES DERIVADAS NO PONDERADAS
# ============================================================

def construir_variables_derivadas(
    matriz
):

    resultado = matriz.copy()

    # --------------------------------------------------------
    # DEMANDA / HOSPEDAJE
    # --------------------------------------------------------

    if (
        "Demanda_Visitantes_Sitios"
        in resultado.columns
        and
        "Hosp_Arribos"
        in resultado.columns
    ):

        resultado[
            "Deriv_Visitantes_Sitios_por_1000_Arribos"
        ] = (
            resultado[
                "Demanda_Visitantes_Sitios"
            ]
            /
            resultado[
                "Hosp_Arribos"
            ].replace(
                0,
                np.nan
            )
            * 1000
        )

    # --------------------------------------------------------
    # INTENSIDAD EMPRESARIAL
    # --------------------------------------------------------

    if (
        "Oferta_Prestadores_Formales_Total"
        in resultado.columns
        and
        "Hosp_Arribos"
        in resultado.columns
    ):

        resultado[
            "Deriv_Prestadores_por_100mil_Arribos"
        ] = (
            resultado[
                "Oferta_Prestadores_Formales_Total"
            ]
            /
            resultado[
                "Hosp_Arribos"
            ].replace(
                0,
                np.nan
            )
            * 100000
        )

    # --------------------------------------------------------
    # CONECTIVIDAD POR ARRIBOS
    # --------------------------------------------------------

    if (
        "Aereo_Pasajeros_Total"
        in resultado.columns
        and
        "Hosp_Arribos"
        in resultado.columns
    ):

        resultado[
            "Deriv_Pasajeros_Aereos_por_Arribo"
        ] = (
            resultado[
                "Aereo_Pasajeros_Total"
            ]
            /
            resultado[
                "Hosp_Arribos"
            ].replace(
                0,
                np.nan
            )
        )

    # --------------------------------------------------------
    # CAPITAL NATURAL POR PRESTADOR
    # --------------------------------------------------------

    if (
        "Capital_ANP_Total"
        in resultado.columns
        and
        "Oferta_Prestadores_Formales_Total"
        in resultado.columns
    ):

        resultado[
            "Deriv_ANP_por_100_Prestadores"
        ] = (
            resultado[
                "Capital_ANP_Total"
            ]
            /
            resultado[
                "Oferta_Prestadores_Formales_Total"
            ].replace(
                0,
                np.nan
            )
            * 100
        )

    # --------------------------------------------------------
    # SUPERFICIE NATURAL POR PRESTADOR
    # --------------------------------------------------------

    if (
        "Capital_Superficie_Atribuida_Referencial_Ha"
        in resultado.columns
        and
        "Oferta_Prestadores_Formales_Total"
        in resultado.columns
    ):

        resultado[
            "Deriv_Ha_Protegidas_por_Prestador"
        ] = (
            resultado[
                "Capital_Superficie_Atribuida_Referencial_Ha"
            ]
            /
            resultado[
                "Oferta_Prestadores_Formales_Total"
            ].replace(
                0,
                np.nan
            )
        )

    return resultado


# ============================================================
# MATRIZ REDUCIDA PARA ANÁLISIS
# ============================================================

def construir_matriz_analitica(
    matriz
):

    columnas = [
        "Departamento",

        "Base_IRNA_Estructural",
        "Base_IRNA_Ejecucion",
        "Base_Brecha_IRNA",

        "Demanda_Visitantes_Sitios",
        "Demanda_Participacion_Extranjera_Porcentaje",

        "Hosp_Arribos",
        "Hosp_Pernoctaciones",
        "Hosp_Permanencia_Promedio",
        "Hosp_Permanencia_Extranjeros",
        "Hosp_TNOH_Promedio",

        "Aereo_Pasajeros_Total",
        "Aereo_Pasajeros_Internacionales",
        "Aereo_Participacion_Internacional_Porcentaje",
        "Aereo_Aeropuertos_Con_Registro",

        "Oferta_Hospedajes_Formales",
        "Oferta_Agencias_Formales",
        "Oferta_Restaurantes_Calificados",
        "Oferta_Prestadores_Formales_Total",

        "Capital_ANP_Total",
        "Capital_ANP_Nacional",
        "Capital_ACR_Total",
        "Capital_ACP_Total",
        "Capital_Zonas_Reservadas",
        "Capital_Diversidad_Categorias",
        "Capital_Superficie_Atribuida_Referencial_Ha",

        "Deriv_Visitantes_Sitios_por_1000_Arribos",
        "Deriv_Prestadores_por_100mil_Arribos",
        "Deriv_Pasajeros_Aereos_por_Arribo",
        "Deriv_ANP_por_100_Prestadores",
        "Deriv_Ha_Protegidas_por_Prestador"
    ]

    columnas = seleccionar_columnas_existentes(
        matriz,
        columnas
    )

    return matriz[
        columnas
    ].copy()


# ============================================================
# CONTROL DE DUPLICADOS TERRITORIALES
# ============================================================

def validar_matriz(
    matriz
):

    print(
        "\n" + "=" * 100
    )

    print(
        "CONTROL DE INTEGRIDAD"
    )

    print(
        "=" * 100
    )

    duplicados = matriz[
        "Departamento"
    ].duplicated().sum()

    faltantes_catalogo = (
        set(
            DEPARTAMENTOS_PERU
        )
        - set(
            matriz[
                "Departamento"
            ]
        )
    )

    fuera_catalogo = (
        set(
            matriz[
                "Departamento"
            ]
        )
        - set(
            DEPARTAMENTOS_PERU
        )
    )

    print(
        f"Filas matriz              : "
        f"{len(matriz)}"
    )

    print(
        f"Departamentos únicos      : "
        f"{matriz['Departamento'].nunique()}"
    )

    print(
        f"Duplicados territoriales  : "
        f"{duplicados}"
    )

    print(
        f"Faltantes del catálogo    : "
        f"{len(faltantes_catalogo)}"
    )

    print(
        f"Fuera del catálogo        : "
        f"{len(fuera_catalogo)}"
    )

    if faltantes_catalogo:

        print(
            "\nFALTANTES:"
        )

        for x in sorted(
            faltantes_catalogo
        ):

            print(
                f" - {x}"
            )

    if fuera_catalogo:

        print(
            "\nFUERA DEL CATÁLOGO:"
        )

        for x in sorted(
            fuera_catalogo
        ):

            print(
                f" - {x}"
            )

    if (
        len(matriz) == 25
        and
        matriz[
            "Departamento"
        ].nunique() == 25
        and
        duplicados == 0
        and
        not faltantes_catalogo
        and
        not fuera_catalogo
    ):

        print(
            "\n✓ MATRIZ TERRITORIAL ÍNTEGRA"
        )

    else:

        print(
            "\n⚠ REVISAR INTEGRIDAD TERRITORIAL"
        )


# ============================================================
# FORMATO EXCEL
# ============================================================

def formatear_excel(ruta):

    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(
        ruta
    )

    for ws in wb.worksheets:

        ws.freeze_panes = "A2"

        if ws.max_row > 1:

            ws.auto_filter.ref = (
                ws.dimensions
            )

        for celda in ws[1]:

            celda.font = Font(
                bold=True
            )

            celda.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        for columna in range(
            1,
            ws.max_column + 1
        ):

            letra = get_column_letter(
                columna
            )

            ancho = 0

            for fila in range(
                1,
                min(
                    ws.max_row,
                    200
                ) + 1
            ):

                valor = ws.cell(
                    fila,
                    columna
                ).value

                if valor is not None:

                    ancho = max(
                        ancho,
                        len(
                            str(valor)
                        )
                    )

            ws.column_dimensions[
                letra
            ].width = min(
                max(
                    ancho + 2,
                    12
                ),
                38
            )

    wb.save(
        ruta
    )


# ============================================================
# MAIN
# ============================================================

def main():

    print(
        "=" * 100
    )

    print(
        "RADAR TURISMO NATURALEZA "
        "Y AVENTURA - PERÚ"
    )

    print(
        "23 - INTEGRADOR FASE 2"
    )

    print(
        "=" * 100
    )

    matriz = construir_matriz_maestra()

    validar_matriz(
        matriz
    )

    matriz = construir_variables_derivadas(
        matriz
    )

    cobertura = construir_control_cobertura(
        matriz
    )

    analitica = construir_matriz_analitica(
        matriz
    )

    # ========================================================
    # CONTROL GLOBAL
    # ========================================================

    print(
        "\n" + "=" * 100
    )

    print(
        "COBERTURA POR CAPA"
    )

    print(
        "=" * 100
    )

    print(
        cobertura.to_string(
            index=False,
            formatters={
                "Cobertura_Porcentaje":
                    lambda x:
                    f"{x:.1f}%"
            }
        )
    )

    print(
        "\n" + "=" * 100
    )

    print(
        "MATRIZ ANALÍTICA - PRIMERAS 10 FILAS"
    )

    print(
        "=" * 100
    )

    print(
        analitica
        .head(10)
        .to_string(
            index=False
        )
    )

    # ========================================================
    # EXPORTAR
    # ========================================================

    with pd.ExcelWriter(
        SALIDA,
        engine="openpyxl"
    ) as writer:

        matriz.to_excel(
            writer,
            sheet_name="01_Matriz_Maestra",
            index=False
        )

        analitica.to_excel(
            writer,
            sheet_name="02_Matriz_Analitica",
            index=False
        )

        cobertura.to_excel(
            writer,
            sheet_name="03_Control_Cobertura",
            index=False
        )

        fuentes = pd.DataFrame(
            {
                "Capa": [
                    "IRNA base",
                    "Demanda",
                    "Hospedaje",
                    "Conectividad aérea",
                    "Oferta formal",
                    "Capital natural"
                ],

                "Archivo": [
                    str(
                        ARCHIVO_IRNA
                    ),
                    str(
                        ARCHIVO_DEMANDA
                    ),
                    str(
                        ARCHIVO_HOSPEDAJE
                    ),
                    str(
                        ARCHIVO_AEREO
                    ),
                    str(
                        ARCHIVO_OFERTA
                    ),
                    str(
                        ARCHIVO_CAPITAL
                    )
                ]
            }
        )

        fuentes.to_excel(
            writer,
            sheet_name="04_Fuentes",
            index=False
        )

        metodologia = pd.DataFrame(
            {
                "Tema": [
                    "Objetivo",
                    "Unidad territorial",
                    "Cobertura",
                    "Índice compuesto",
                    "Variables derivadas",
                    "Capital natural",
                    "Siguiente etapa"
                ],

                "Criterio": [
                    "Integrar las capas validadas de la Fase 2 en una sola matriz.",
                    "25 territorios oficiales del Radar.",
                    "Se conserva explícitamente la ausencia de registro y los valores cero.",
                    "Esta etapa NO calcula aún el IRNA-C.",
                    "Las variables derivadas son descriptivas y no incluyen ponderaciones.",
                    "La superficie ANP territorial es referencial hasta contar con intersección GIS exacta.",
                    "La siguiente etapa analizará normalización, correlaciones y diseño del IRNA-C."
                ]
            }
        )

        metodologia.to_excel(
            writer,
            sheet_name="05_Metodologia",
            index=False
        )

    formatear_excel(
        SALIDA
    )

    # ========================================================
    # CONTROL FINAL
    # ========================================================

    print(
        "\n" + "=" * 100
    )

    print(
        "CONTROL FINAL"
    )

    print(
        "=" * 100
    )

    print(
        f"TERRITORIOS MATRIZ       : "
        f"{len(matriz)}"
    )

    print(
        f"COLUMNAS MATRIZ          : "
        f"{len(matriz.columns)}"
    )

    print(
        f"COLUMNAS ANALÍTICAS      : "
        f"{len(analitica.columns)}"
    )

    print(
        f"CAPAS INTEGRADAS         : "
        f"6"
    )

    print(
        "\nARCHIVO GENERADO:"
    )

    print(
        SALIDA
    )

    print(
        "\n✓ MATRIZ MAESTRA "
        "DE FASE 2 COMPLETADA"
    )

    print(
        "FIN"
    )


if __name__ == "__main__":
    main()