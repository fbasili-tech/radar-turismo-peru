from pathlib import Path

import numpy as np
import pandas as pd


# ============================================================
# RADAR TURISMO NATURALEZA Y AVENTURA - PERÚ
#
# 26 - IRNA-C
# MODELO BASE + ANÁLISIS DE SENSIBILIDAD
#
# OBJETIVOS
# ------------------------------------------------------------
# 1. Construir el primer IRNA-C experimental.
# 2. Asignar pesos iguales entre las 5 dimensiones.
# 3. Calcular scores por dimensión.
# 4. Generar ranking nacional.
# 5. Probar escenarios alternativos de ponderación.
# 6. Medir estabilidad del ranking.
# 7. Comparar IRNA-C con IRNA estructural y ejecución.
#
# IMPORTANTE:
# Este modelo sigue siendo experimental.
# El análisis de sensibilidad determinará si el ranking
# es suficientemente robusto antes de considerarlo definitivo.
# ============================================================


# ============================================================
# ARCHIVOS
# ============================================================

ENTRADA_NORMALIZADOS = Path(
    "outputs/radar_fase2_indicadores_normalizados_irnac_2026.xlsx"
)

ENTRADA_MAESTRA = Path(
    "outputs/radar_fase2_matriz_maestra_2026.xlsx"
)

SALIDA = Path(
    "outputs/radar_fase2_irnac_modelo_base_2026.xlsx"
)

SALIDA.parent.mkdir(
    parents=True,
    exist_ok=True
)


# ============================================================
# HOJAS
# ============================================================

HOJA_NORMALIZADOS = "03_Normalizados_IRNAC"
HOJA_CATALOGO = "02_Catalogo_Indicadores"
HOJA_MAESTRA = "01_Matriz_Maestra"


# ============================================================
# PESOS DEL MODELO BASE
# ============================================================

PESOS_BASE = {
    "DEMANDA": 0.20,
    "DESEMPENO_TURISTICO": 0.20,
    "CONECTIVIDAD": 0.20,
    "OFERTA_FORMAL": 0.20,
    "CAPITAL_NATURAL": 0.20
}


# ============================================================
# ESCENARIOS DE SENSIBILIDAD
# ============================================================

ESCENARIOS = {

    "BASE_20_20_20_20_20": {
        "DEMANDA": 0.20,
        "DESEMPENO_TURISTICO": 0.20,
        "CONECTIVIDAD": 0.20,
        "OFERTA_FORMAL": 0.20,
        "CAPITAL_NATURAL": 0.20
    },

    "CAPITAL_NATURAL_ALTO": {
        "DEMANDA": 0.15,
        "DESEMPENO_TURISTICO": 0.15,
        "CONECTIVIDAD": 0.15,
        "OFERTA_FORMAL": 0.15,
        "CAPITAL_NATURAL": 0.40
    },

    "MERCADO_ALTO": {
        "DEMANDA": 0.30,
        "DESEMPENO_TURISTICO": 0.25,
        "CONECTIVIDAD": 0.15,
        "OFERTA_FORMAL": 0.15,
        "CAPITAL_NATURAL": 0.15
    },

    "CONECTIVIDAD_ALTA": {
        "DEMANDA": 0.15,
        "DESEMPENO_TURISTICO": 0.15,
        "CONECTIVIDAD": 0.40,
        "OFERTA_FORMAL": 0.15,
        "CAPITAL_NATURAL": 0.15
    },

    "OFERTA_ALTA": {
        "DEMANDA": 0.15,
        "DESEMPENO_TURISTICO": 0.15,
        "CONECTIVIDAD": 0.15,
        "OFERTA_FORMAL": 0.40,
        "CAPITAL_NATURAL": 0.15
    },

    "DEMANDA_CAPITAL": {
        "DEMANDA": 0.30,
        "DESEMPENO_TURISTICO": 0.15,
        "CONECTIVIDAD": 0.10,
        "OFERTA_FORMAL": 0.15,
        "CAPITAL_NATURAL": 0.30
    }
}


# ============================================================
# UTILIDADES
# ============================================================

def validar_archivo(ruta):

    if not ruta.exists():

        raise FileNotFoundError(
            f"No existe el archivo requerido:\n{ruta}"
        )


def validar_pesos(pesos):

    total = sum(
        pesos.values()
    )

    if not np.isclose(
        total,
        1.0
    ):

        raise ValueError(
            f"Los pesos suman {total:.6f} "
            "y deberían sumar 1.0"
        )


# ============================================================
# CARGAR DATOS
# ============================================================

def cargar_datos():

    validar_archivo(
        ENTRADA_NORMALIZADOS
    )

    validar_archivo(
        ENTRADA_MAESTRA
    )

    print(
        "\nCargando indicadores normalizados..."
    )

    normalizados = pd.read_excel(
        ENTRADA_NORMALIZADOS,
        sheet_name=HOJA_NORMALIZADOS
    )

    catalogo = pd.read_excel(
        ENTRADA_NORMALIZADOS,
        sheet_name=HOJA_CATALOGO
    )

    print(
        "\nCargando matriz maestra..."
    )

    maestra = pd.read_excel(
        ENTRADA_MAESTRA,
        sheet_name=HOJA_MAESTRA
    )

    print(
        f"Territorios normalizados : "
        f"{len(normalizados)}"
    )

    print(
        f"Indicadores normalizados : "
        f"{len(normalizados.columns) - 1}"
    )

    print(
        f"Variables catálogo       : "
        f"{len(catalogo)}"
    )

    return (
        normalizados,
        catalogo,
        maestra
    )


# ============================================================
# OBTENER INDICADORES POR DIMENSIÓN
# ============================================================

def construir_arquitectura(
    catalogo,
    normalizados
):

    incluidos = catalogo[
        catalogo[
            "Decision"
        ] == "INCLUIR"
    ].copy()

    filas = []

    for dimension, grupo in incluidos.groupby(
        "Dimension"
    ):

        indicadores = []

        for variable in grupo[
            "Variable"
        ]:

            columna = (
                "N_"
                + str(variable)
            )

            if columna in normalizados.columns:

                indicadores.append(
                    columna
                )

        cantidad = len(
            indicadores
        )

        if cantidad == 0:

            continue

        peso_dimension = PESOS_BASE[
            dimension
        ]

        peso_indicador = (
            peso_dimension
            / cantidad
        )

        for indicador in indicadores:

            filas.append(
                {
                    "Dimension":
                        dimension,

                    "Indicador_Normalizado":
                        indicador,

                    "Peso_Dimension_Base":
                        peso_dimension,

                    "Indicadores_Dimension":
                        cantidad,

                    "Peso_Indicador_Base":
                        peso_indicador
                }
            )

    arquitectura = pd.DataFrame(
        filas
    )

    return arquitectura


# ============================================================
# SCORE POR DIMENSIÓN
# ============================================================

def calcular_scores_dimension(
    normalizados,
    arquitectura
):

    resultado = pd.DataFrame(
        {
            "Departamento":
                normalizados[
                    "Departamento"
                ]
        }
    )

    dimensiones = (
        arquitectura[
            "Dimension"
        ]
        .unique()
        .tolist()
    )

    for dimension in dimensiones:

        indicadores = arquitectura.loc[
            arquitectura[
                "Dimension"
            ] == dimension,
            "Indicador_Normalizado"
        ].tolist()

        resultado[
            "Score_"
            + dimension
        ] = (
            normalizados[
                indicadores
            ]
            .mean(
                axis=1
            )
        )

    return resultado


# ============================================================
# CALCULAR IRNA-C
# ============================================================

def calcular_indice(
    scores,
    pesos,
    nombre_columna
):

    validar_pesos(
        pesos
    )

    indice = pd.Series(
        0.0,
        index=scores.index
    )

    for dimension, peso in pesos.items():

        columna = (
            "Score_"
            + dimension
        )

        if columna not in scores.columns:

            raise ValueError(
                f"No existe {columna}"
            )

        indice = (
            indice
            +
            scores[
                columna
            ]
            * peso
        )

    resultado = scores.copy()

    resultado[
        nombre_columna
    ] = indice

    return resultado


# ============================================================
# CATEGORIZAR IRNA-C
# ============================================================

def categorizar_irnac(valor):

    if pd.isna(valor):

        return "SIN DATO"

    if valor >= 75:

        return "LÍDER"

    if valor >= 60:

        return "ALTA COMPETITIVIDAD"

    if valor >= 45:

        return "EN CONSOLIDACIÓN"

    if valor >= 30:

        return "EMERGENTE"

    return "REZAGADO"


# ============================================================
# RANKING BASE
# ============================================================

def construir_ranking_base(
    scores
):

    resultado = calcular_indice(
        scores,
        PESOS_BASE,
        "IRNA_C"
    )

    resultado[
        "Categoria_IRNA_C"
    ] = (
        resultado[
            "IRNA_C"
        ]
        .apply(
            categorizar_irnac
        )
    )

    resultado = (
        resultado
        .sort_values(
            "IRNA_C",
            ascending=False
        )
        .reset_index(
            drop=True
        )
    )

    resultado.insert(
        0,
        "Ranking_IRNA_C",
        range(
            1,
            len(resultado) + 1
        )
    )

    return resultado


# ============================================================
# ESCENARIOS
# ============================================================

def construir_escenarios(
    scores
):

    resultado = pd.DataFrame(
        {
            "Departamento":
                scores[
                    "Departamento"
                ]
        }
    )

    for nombre, pesos in (
        ESCENARIOS.items()
    ):

        validar_pesos(
            pesos
        )

        score = pd.Series(
            0.0,
            index=scores.index
        )

        for dimension, peso in pesos.items():

            score = (
                score
                +
                scores[
                    "Score_"
                    + dimension
                ]
                * peso
            )

        resultado[
            "IRNAC_"
            + nombre
        ] = score

    return resultado


# ============================================================
# RANKINGS DE ESCENARIOS
# ============================================================

def construir_rankings_escenarios(
    escenarios
):

    resultado = pd.DataFrame(
        {
            "Departamento":
                escenarios[
                    "Departamento"
                ]
        }
    )

    columnas_scores = [
        c
        for c in escenarios.columns
        if c.startswith(
            "IRNAC_"
        )
    ]

    for columna in columnas_scores:

        resultado[
            columna.replace(
                "IRNAC_",
                "Ranking_"
            )
        ] = (
            escenarios[
                columna
            ]
            .rank(
                method="min",
                ascending=False
            )
            .astype(int)
        )

    return resultado


# ============================================================
# SENSIBILIDAD TERRITORIAL
# ============================================================

def construir_sensibilidad(
    escenarios,
    rankings
):

    columnas_score = [
        c
        for c in escenarios.columns
        if c.startswith(
            "IRNAC_"
        )
    ]

    columnas_rank = [
        c
        for c in rankings.columns
        if c.startswith(
            "Ranking_"
        )
    ]

    filas = []

    for i, departamento in enumerate(
        escenarios[
            "Departamento"
        ]
    ):

        valores = pd.to_numeric(
            escenarios.loc[
                i,
                columnas_score
            ],
            errors="coerce"
        )

        posiciones = pd.to_numeric(
            rankings.loc[
                i,
                columnas_rank
            ],
            errors="coerce"
        )

        ranking_base_col = (
            "Ranking_BASE_20_20_20_20_20"
        )

        ranking_base = rankings.loc[
            i,
            ranking_base_col
        ]

        desviaciones = abs(
            posiciones
            - ranking_base
        )

        filas.append(
            {
                "Departamento":
                    departamento,

                "IRNAC_Min":
                    valores.min(),

                "IRNAC_Max":
                    valores.max(),

                "Rango_IRNAC":
                    (
                        valores.max()
                        - valores.min()
                    ),

                "IRNAC_Promedio":
                    valores.mean(),

                "Desviacion_IRNAC":
                    valores.std(),

                "Ranking_Base":
                    ranking_base,

                "Ranking_Mejor":
                    posiciones.min(),

                "Ranking_Peor":
                    posiciones.max(),

                "Rango_Ranking":
                    (
                        posiciones.max()
                        - posiciones.min()
                    ),

                "Cambio_Maximo_vs_Base":
                    desviaciones.max()
            }
        )

    sensibilidad = pd.DataFrame(
        filas
    )

    return sensibilidad


# ============================================================
# CLASIFICAR ROBUSTEZ
# ============================================================

def clasificar_robustez(
    rango_ranking
):

    if pd.isna(
        rango_ranking
    ):

        return "SIN DATO"

    if rango_ranking <= 2:

        return "MUY ROBUSTO"

    if rango_ranking <= 4:

        return "ROBUSTO"

    if rango_ranking <= 7:

        return "SENSIBLE"

    return "MUY SENSIBLE"


# ============================================================
# CORRELACIÓN ENTRE RANKINGS
# ============================================================

def construir_correlacion_rankings(
    rankings
):

    columnas = [
        c
        for c in rankings.columns
        if c.startswith(
            "Ranking_"
        )
    ]

    matriz = rankings[
        columnas
    ].corr(
        method="spearman"
    )

    return matriz


# ============================================================
# RESUMEN ESCENARIOS
# ============================================================

def construir_resumen_escenarios():

    filas = []

    for nombre, pesos in (
        ESCENARIOS.items()
    ):

        filas.append(
            {
                "Escenario":
                    nombre,

                "Peso_Demanda":
                    pesos[
                        "DEMANDA"
                    ],

                "Peso_Desempeno":
                    pesos[
                        "DESEMPENO_TURISTICO"
                    ],

                "Peso_Conectividad":
                    pesos[
                        "CONECTIVIDAD"
                    ],

                "Peso_Oferta_Formal":
                    pesos[
                        "OFERTA_FORMAL"
                    ],

                "Peso_Capital_Natural":
                    pesos[
                        "CAPITAL_NATURAL"
                    ],

                "Suma_Pesos":
                    sum(
                        pesos.values()
                    )
            }
        )

    return pd.DataFrame(
        filas
    )


# ============================================================
# INTEGRAR IRNA ANTERIOR
# ============================================================

def integrar_irna_previo(
    ranking,
    maestra
):

    columnas_posibles = [
        "Departamento",
        "Base_IRNA_Estructural",
        "Base_IRNA_Ejecucion",
        "Base_Brecha_IRNA"
    ]

    columnas = [
        c
        for c in columnas_posibles
        if c in maestra.columns
    ]

    contraste = maestra[
        columnas
    ].copy()

    resultado = ranking.merge(
        contraste,
        on="Departamento",
        how="left"
    )

    if (
        "Base_IRNA_Estructural"
        in resultado.columns
    ):

        resultado[
            "Brecha_IRNAC_vs_Estructural"
        ] = (
            resultado[
                "IRNA_C"
            ]
            -
            resultado[
                "Base_IRNA_Estructural"
            ]
        )

    if (
        "Base_IRNA_Ejecucion"
        in resultado.columns
    ):

        resultado[
            "Brecha_IRNAC_vs_Ejecucion"
        ] = (
            resultado[
                "IRNA_C"
            ]
            -
            resultado[
                "Base_IRNA_Ejecucion"
            ]
        )

    return resultado


# ============================================================
# CUADRANTE ESTRATÉGICO
# ============================================================

def asignar_cuadrante(
    fila
):

    irnac = fila.get(
        "IRNA_C",
        np.nan
    )

    estructural = fila.get(
        "Base_IRNA_Estructural",
        np.nan
    )

    if (
        pd.isna(irnac)
        or pd.isna(estructural)
    ):

        return "SIN CLASIFICAR"

    if (
        estructural >= 60
        and irnac >= 60
    ):

        return (
            "Q1 - ALTO POTENCIAL / "
            "ALTA COMPETITIVIDAD"
        )

    if (
        estructural >= 60
        and irnac < 60
    ):

        return (
            "Q2 - ALTO POTENCIAL / "
            "COMPETITIVIDAD POR DESARROLLAR"
        )

    if (
        estructural < 60
        and irnac >= 60
    ):

        return (
            "Q3 - COMPETITIVIDAD SUPERIOR "
            "AL POTENCIAL MEDIDO"
        )

    return (
        "Q4 - DESARROLLO EMERGENTE"
    )


# ============================================================
# RESUMEN NACIONAL
# ============================================================

def construir_resumen_nacional(
    ranking,
    sensibilidad
):

    combinado = ranking[
        [
            "Departamento",
            "IRNA_C"
        ]
    ].merge(
        sensibilidad,
        on="Departamento",
        how="left"
    )

    resumen = pd.DataFrame(
        {
            "Indicador": [
                "Territorios analizados",
                "IRNA-C promedio",
                "IRNA-C mediana",
                "IRNA-C máximo",
                "IRNA-C mínimo",
                "Territorios líder",
                "Alta competitividad",
                "En consolidación",
                "Emergentes",
                "Rezagados",
                "Rango ranking promedio en sensibilidad",
                "Cambio máximo promedio vs base"
            ],

            "Valor": [
                len(
                    ranking
                ),

                ranking[
                    "IRNA_C"
                ].mean(),

                ranking[
                    "IRNA_C"
                ].median(),

                ranking[
                    "IRNA_C"
                ].max(),

                ranking[
                    "IRNA_C"
                ].min(),

                (
                    ranking[
                        "Categoria_IRNA_C"
                    ]
                    == "LÍDER"
                ).sum(),

                (
                    ranking[
                        "Categoria_IRNA_C"
                    ]
                    == "ALTA COMPETITIVIDAD"
                ).sum(),

                (
                    ranking[
                        "Categoria_IRNA_C"
                    ]
                    == "EN CONSOLIDACIÓN"
                ).sum(),

                (
                    ranking[
                        "Categoria_IRNA_C"
                    ]
                    == "EMERGENTE"
                ).sum(),

                (
                    ranking[
                        "Categoria_IRNA_C"
                    ]
                    == "REZAGADO"
                ).sum(),

                combinado[
                    "Rango_Ranking"
                ].mean(),

                combinado[
                    "Cambio_Maximo_vs_Base"
                ].mean()
            ]
        }
    )

    return resumen


# ============================================================
# FORMATO EXCEL
# ============================================================

def formatear_excel(
    ruta
):

    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(
        ruta
    )

    for ws in wb.worksheets:

        ws.freeze_panes = "A2"

        if (
            ws.max_row > 1
            and ws.max_column > 1
        ):

            ws.auto_filter.ref = (
                ws.dimensions
            )

        for celda in ws[1]:

            celda.font = Font(
                bold=True
            )

            celda.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        for columna in range(
            1,
            ws.max_column + 1
        ):

            letra = get_column_letter(
                columna
            )

            ancho = 0

            for fila in range(
                1,
                min(
                    ws.max_row,
                    200
                ) + 1
            ):

                valor = ws.cell(
                    fila,
                    columna
                ).value

                if valor is not None:

                    ancho = max(
                        ancho,
                        len(
                            str(valor)
                        )
                    )

            ws.column_dimensions[
                letra
            ].width = min(
                max(
                    ancho + 2,
                    12
                ),
                45
            )

    wb.save(
        ruta
    )


# ============================================================
# MAIN
# ============================================================

def main():

    print(
        "=" * 100
    )

    print(
        "RADAR TURISMO NATURALEZA "
        "Y AVENTURA - PERÚ"
    )

    print(
        "26 - IRNA-C MODELO BASE "
        "Y ANÁLISIS DE SENSIBILIDAD"
    )

    print(
        "=" * 100
    )

    # --------------------------------------------------------
    # VALIDAR PESOS
    # --------------------------------------------------------

    validar_pesos(
        PESOS_BASE
    )

    for pesos in (
        ESCENARIOS.values()
    ):

        validar_pesos(
            pesos
        )

    # --------------------------------------------------------
    # CARGA
    # --------------------------------------------------------

    (
        normalizados,
        catalogo,
        maestra
    ) = cargar_datos()

    # --------------------------------------------------------
    # ARQUITECTURA
    # --------------------------------------------------------

    arquitectura = construir_arquitectura(
        catalogo,
        normalizados
    )

    # --------------------------------------------------------
    # SCORES DIMENSIONALES
    # --------------------------------------------------------

    scores = calcular_scores_dimension(
        normalizados,
        arquitectura
    )

    # --------------------------------------------------------
    # RANKING BASE
    # --------------------------------------------------------

    ranking = construir_ranking_base(
        scores
    )

    # --------------------------------------------------------
    # ESCENARIOS
    # --------------------------------------------------------

    escenarios = construir_escenarios(
        scores
    )

    rankings_escenarios = (
        construir_rankings_escenarios(
            escenarios
        )
    )

    # --------------------------------------------------------
    # SENSIBILIDAD
    # --------------------------------------------------------

    sensibilidad = construir_sensibilidad(
        escenarios,
        rankings_escenarios
    )

    sensibilidad[
        "Robustez_Ranking"
    ] = (
        sensibilidad[
            "Rango_Ranking"
        ]
        .apply(
            clasificar_robustez
        )
    )

    # --------------------------------------------------------
    # CORRELACIÓN DE RANKINGS
    # --------------------------------------------------------

    correlacion_rankings = (
        construir_correlacion_rankings(
            rankings_escenarios
        )
    )

    # --------------------------------------------------------
    # CONTRASTE IRNA PREVIO
    # --------------------------------------------------------

    ranking_contraste = (
        integrar_irna_previo(
            ranking,
            maestra
        )
    )

    ranking_contraste[
        "Cuadrante_Estrategico"
    ] = (
        ranking_contraste
        .apply(
            asignar_cuadrante,
            axis=1
        )
    )

    # --------------------------------------------------------
    # RESÚMENES
    # --------------------------------------------------------

    resumen_escenarios = (
        construir_resumen_escenarios()
    )

    resumen_nacional = (
        construir_resumen_nacional(
            ranking,
            sensibilidad
        )
    )

    # ========================================================
    # CONSOLA
    # ========================================================

    print(
        "\n" + "=" * 100
    )

    print(
        "ARQUITECTURA DEL MODELO BASE"
    )

    print(
        "=" * 100
    )

    print(
        arquitectura.to_string(
            index=False,
            formatters={
                "Peso_Dimension_Base":
                    lambda x:
                    f"{x:.1%}",

                "Peso_Indicador_Base":
                    lambda x:
                    f"{x:.2%}"
            }
        )
    )

    print(
        "\n" + "=" * 100
    )

    print(
        "RANKING IRNA-C - MODELO BASE"
    )

    print(
        "=" * 100
    )

    columnas_ranking = [
        "Ranking_IRNA_C",
        "Departamento",
        "IRNA_C",
        "Categoria_IRNA_C",
        "Score_DEMANDA",
        "Score_DESEMPENO_TURISTICO",
        "Score_CONECTIVIDAD",
        "Score_OFERTA_FORMAL",
        "Score_CAPITAL_NATURAL"
    ]

    print(
        ranking[
            columnas_ranking
        ]
        .to_string(
            index=False,
            formatters={
                "IRNA_C":
                    lambda x:
                    f"{x:.1f}",

                "Score_DEMANDA":
                    lambda x:
                    f"{x:.1f}",

                "Score_DESEMPENO_TURISTICO":
                    lambda x:
                    f"{x:.1f}",

                "Score_CONECTIVIDAD":
                    lambda x:
                    f"{x:.1f}",

                "Score_OFERTA_FORMAL":
                    lambda x:
                    f"{x:.1f}",

                "Score_CAPITAL_NATURAL":
                    lambda x:
                    f"{x:.1f}"
            }
        )
    )

    print(
        "\n" + "=" * 100
    )

    print(
        "SENSIBILIDAD DEL RANKING"
    )

    print(
        "=" * 100
    )

    sensibilidad_impresion = (
        sensibilidad
        .sort_values(
            [
                "Rango_Ranking",
                "Ranking_Base"
            ],
            ascending=[
                False,
                True
            ]
        )
    )

    print(
        sensibilidad_impresion[
            [
                "Departamento",
                "Ranking_Base",
                "Ranking_Mejor",
                "Ranking_Peor",
                "Rango_Ranking",
                "Cambio_Maximo_vs_Base",
                "Robustez_Ranking"
            ]
        ]
        .to_string(
            index=False
        )
    )

    print(
        "\n" + "=" * 100
    )

    print(
        "CORRELACIÓN ENTRE ESCENARIOS"
    )

    print(
        "=" * 100
    )

    print(
        correlacion_rankings.round(
            3
        ).to_string()
    )

    # --------------------------------------------------------
    # CONTROL
    # --------------------------------------------------------

    n_robustos = int(
        sensibilidad[
            "Robustez_Ranking"
        ]
        .isin(
            [
                "MUY ROBUSTO",
                "ROBUSTO"
            ]
        )
        .sum()
    )

    n_sensibles = (
        len(
            sensibilidad
        )
        - n_robustos
    )

    minimo_correlacion = (
        correlacion_rankings
        .where(
            ~np.eye(
                len(
                    correlacion_rankings
                ),
                dtype=bool
            )
        )
        .min()
        .min()
    )

    print(
        "\n" + "=" * 100
    )

    print(
        "CONTROL FINAL"
    )

    print(
        "=" * 100
    )

    print(
        f"TERRITORIOS                : "
        f"{len(ranking)}"
    )

    print(
        f"DIMENSIONES                : "
        f"{len(PESOS_BASE)}"
    )

    print(
        f"INDICADORES                : "
        f"{len(arquitectura)}"
    )

    print(
        f"ESCENARIOS ANALIZADOS      : "
        f"{len(ESCENARIOS)}"
    )

    print(
        f"TERRITORIOS ROBUSTOS       : "
        f"{n_robustos}"
    )

    print(
        f"TERRITORIOS SENSIBLES      : "
        f"{n_sensibles}"
    )

    print(
        f"CORRELACIÓN MÍNIMA RANKING : "
        f"{minimo_correlacion:.3f}"
    )

    print(
        f"IRNA-C MÁXIMO              : "
        f"{ranking['IRNA_C'].max():.2f}"
    )

    print(
        f"IRNA-C MEDIANA             : "
        f"{ranking['IRNA_C'].median():.2f}"
    )

    print(
        f"IRNA-C MÍNIMO              : "
        f"{ranking['IRNA_C'].min():.2f}"
    )

    # ========================================================
    # EXPORTAR
    # ========================================================

    with pd.ExcelWriter(
        SALIDA,
        engine="openpyxl"
    ) as writer:

        resumen_nacional.to_excel(
            writer,
            sheet_name="01_Resumen_Nacional",
            index=False
        )

        ranking_contraste.to_excel(
            writer,
            sheet_name="02_Ranking_IRNAC",
            index=False
        )

        scores.to_excel(
            writer,
            sheet_name="03_Scores_Dimensiones",
            index=False
        )

        arquitectura.to_excel(
            writer,
            sheet_name="04_Arquitectura",
            index=False
        )

        escenarios.to_excel(
            writer,
            sheet_name="05_Escenarios_Scores",
            index=False
        )

        rankings_escenarios.to_excel(
            writer,
            sheet_name="06_Escenarios_Rankings",
            index=False
        )

        sensibilidad.to_excel(
            writer,
            sheet_name="07_Sensibilidad",
            index=False
        )

        correlacion_rankings.to_excel(
            writer,
            sheet_name="08_Correlacion_Rankings"
        )

        resumen_escenarios.to_excel(
            writer,
            sheet_name="09_Pesos_Escenarios",
            index=False
        )

        metodologia = pd.DataFrame(
            {
                "Tema": [
                    "Modelo base",
                    "Peso dimensiones",
                    "Peso indicadores",
                    "Normalización",
                    "Sensibilidad",
                    "Robustez muy alta",
                    "Robustez",
                    "Sensibilidad media",
                    "Sensibilidad alta",
                    "IRNA previo",
                    "Interpretación"
                ],

                "Criterio": [
                    (
                        "El IRNA-C base utiliza cinco "
                        "dimensiones con igual ponderación."
                    ),

                    (
                        "Cada dimensión representa 20% "
                        "del índice total."
                    ),

                    (
                        "Los indicadores se reparten "
                        "equitativamente dentro de su dimensión."
                    ),

                    (
                        "Los indicadores provienen de la "
                        "normalización robusta 0-100 del paso 25."
                    ),

                    (
                        "Se prueban escenarios alternativos "
                        "para medir estabilidad del ranking."
                    ),

                    (
                        "Rango de ranking entre escenarios "
                        "igual o menor a 2 posiciones."
                    ),

                    (
                        "Rango entre 3 y 4 posiciones."
                    ),

                    (
                        "Rango entre 5 y 7 posiciones."
                    ),

                    (
                        "Rango mayor a 7 posiciones."
                    ),

                    (
                        "IRNA estructural y ejecución no entran "
                        "en el cálculo; se utilizan como contraste."
                    ),

                    (
                        "El ranking generado sigue siendo "
                        "experimental hasta revisar sensibilidad."
                    )
                ]
            }
        )

        metodologia.to_excel(
            writer,
            sheet_name="10_Metodologia",
            index=False
        )

    formatear_excel(
        SALIDA
    )

    print(
        "\nARCHIVO GENERADO:"
    )

    print(
        SALIDA
    )

    print(
        "\n✓ MODELO BASE IRNA-C COMPLETADO"
    )

    print(
        "✓ ANÁLISIS DE SENSIBILIDAD COMPLETADO"
    )

    print(
        "✓ RANKING AÚN CON CARÁCTER EXPERIMENTAL"
    )

    print(
        "FIN"
    )


if __name__ == "__main__":
    main()