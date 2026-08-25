from pathlib import Path
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter


# ============================================================
# RADAR TURISMO DE NATURALEZA Y AVENTURA - PERÚ
# ETAPA 14B: DASHBOARD PROFESIONAL
# ============================================================

ENTRADA = Path(
    "outputs/radar_turismo_tablero_base_2026.xlsx"
)

SALIDA = Path(
    "outputs/radar_turismo_dashboard_profesional_2026.xlsx"
)


# ============================================================
# PALETA
# ============================================================

VERDE_OSCURO = "143D33"
VERDE = "2E7D5B"
AZUL = "2F6B9A"
ROJO = "D9534F"
NARANJA = "F0A04B"
AMARILLO = "E4C441"
GRIS_CLARO = "F3F5F7"
GRIS = "D9E1E8"
GRIS_OSCURO = "404040"
BLANCO = "FFFFFF"


# ============================================================
# FUNCIONES
# ============================================================

def color_alerta(valor):
    mapa = {
        "ROJA": ROJO,
        "NARANJA": NARANJA,
        "AMARILLA": AMARILLO,
        "VERDE": VERDE
    }
    return mapa.get(str(valor).upper(), GRIS)


def estilo_encabezado(ws, fila=1):

    for cell in ws[fila]:

        cell.fill = PatternFill(
            "solid",
            fgColor=VERDE_OSCURO
        )

        cell.font = Font(
            color=BLANCO,
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )


def ajustar_columnas(ws):

    for col in ws.columns:

        letra = get_column_letter(
            col[0].column
        )

        max_len = 0

        for cell in col:

            if cell.value is not None:

                max_len = max(
                    max_len,
                    len(str(cell.value))
                )

        ws.column_dimensions[
            letra
        ].width = min(
            max(max_len + 2, 10),
            40
        )


def bloque_kpi(ws, rango, titulo, valor, color):

    ws.merge_cells(rango)

    celda = ws[
        rango.split(":")[0]
    ]

    celda.value = (
        f"{titulo}\n{valor}"
    )

    celda.fill = PatternFill(
        "solid",
        fgColor=color
    )

    celda.font = Font(
        size=14,
        bold=True,
        color=BLANCO
    )

    celda.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )


# ============================================================
# PROCESO PRINCIPAL
# ============================================================

def main():

    print("=" * 100)
    print("RADAR TURISMO NATURALEZA Y AVENTURA - PERÚ")
    print("14B - DASHBOARD PROFESIONAL")
    print("=" * 100)

    ranking = pd.read_excel(
        ENTRADA,
        sheet_name="Ranking_Nacional"
    )

    matriz = pd.read_excel(
        ENTRADA,
        sheet_name="Potencial_vs_Ejecucion"
    )

    alertas = pd.read_excel(
        ENTRADA,
        sheet_name="Alertas_Territoriales"
    )

    oportunidades = pd.read_excel(
        ENTRADA,
        sheet_name="Oportunidades"
    )

    resumen = pd.read_excel(
        ENTRADA,
        sheet_name="Resumen_Nacional"
    )

    print(
        f"\nTerritorios cargados: {len(ranking)}"
    )

    # --------------------------------------------------------
    # MEDIANAS
    # --------------------------------------------------------

    mediana_estructural = ranking[
        "IRNA_Estructural"
    ].median()

    mediana_ejecucion = ranking[
        "IRNA_Ejecucion"
    ].median()

    # --------------------------------------------------------
    # REETIQUETAR CUADRANTES
    # --------------------------------------------------------

    matriz["Cuadrante_IRNA"] = (
        matriz["Cuadrante_IRNA"]
        .replace(
            {
                "Q1 - ALTO POTENCIAL / ALTA EJECUCIÓN":
                    "Q1 - SOBRE MEDIANA EN POTENCIAL Y EJECUCIÓN",

                "Q2 - ALTO POTENCIAL / BAJA EJECUCIÓN":
                    "Q2 - SOBRE MEDIANA EN POTENCIAL / BAJO MEDIANA EN EJECUCIÓN",

                "Q3 - MENOR POTENCIAL / ALTA EJECUCIÓN":
                    "Q3 - BAJO MEDIANA EN POTENCIAL / SOBRE MEDIANA EN EJECUCIÓN",

                "Q4 - MENOR POTENCIAL / BAJA EJECUCIÓN":
                    "Q4 - BAJO MEDIANA EN POTENCIAL Y EJECUCIÓN"
            }
        )
    )

    # --------------------------------------------------------
    # EXPORTAR BASE
    # --------------------------------------------------------

    with pd.ExcelWriter(
        SALIDA,
        engine="openpyxl"
    ) as writer:

        ranking.to_excel(
            writer,
            sheet_name="Ranking",
            index=False
        )

        matriz.to_excel(
            writer,
            sheet_name="Potencial_Ejecucion",
            index=False
        )

        alertas.to_excel(
            writer,
            sheet_name="Alertas",
            index=False
        )

        oportunidades.to_excel(
            writer,
            sheet_name="Oportunidades",
            index=False
        )

        resumen.to_excel(
            writer,
            sheet_name="Indicadores",
            index=False
        )

        pd.DataFrame().to_excel(
            writer,
            sheet_name="Dashboard",
            index=False
        )

    wb = load_workbook(
        SALIDA
    )

    ws = wb[
        "Dashboard"
    ]

    # mover Dashboard al inicio
    wb._sheets.remove(ws)
    wb._sheets.insert(0, ws)

    ws.sheet_view.showGridLines = False

    # ========================================================
    # TÍTULO
    # ========================================================

    ws.merge_cells(
        "A1:P2"
    )

    ws["A1"] = (
        "RADAR DE TURISMO DE NATURALEZA Y AVENTURA DEL PERÚ"
    )

    ws["A1"].font = Font(
        size=22,
        bold=True,
        color=BLANCO
    )

    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=VERDE_OSCURO
    )

    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    ws.merge_cells(
        "A3:P3"
    )

    ws["A3"] = (
        "Inversión pública, fortaleza territorial, ejecución y brechas de gestión | 2026"
    )

    ws["A3"].font = Font(
        italic=True,
        size=11,
        color=GRIS_OSCURO
    )

    ws["A3"].alignment = Alignment(
        horizontal="center"
    )

    # ========================================================
    # KPIs
    # ========================================================

    pim_total = ranking[
        "PIM_Radar"
    ].sum()

    lider = (
        ranking
        .sort_values(
            "IRNA_Estructural",
            ascending=False
        )
        .iloc[0]
    )

    consolidados = ranking[
        ranking["Estado_Territorial"]
        == "FORTALEZA CONSOLIDADA"
    ]

    if len(consolidados) > 0:

        consolidado = (
            consolidados
            .sort_values(
                "IRNA_Estructural",
                ascending=False
            )
            .iloc[0]["Departamento"]
        )

    else:
        consolidado = "—"

    alertas_rojas = (
        ranking["Nivel_Alerta"]
        == "ROJA"
    ).sum()

    bloque_kpi(
        ws,
        "A5:C8",
        "TERRITORIOS ANALIZADOS",
        len(ranking),
        AZUL
    )

    bloque_kpi(
        ws,
        "D5:F8",
        "PIM NÚCLEO RADAR",
        f"S/ {pim_total/1_000_000:.1f} M",
        VERDE
    )

    bloque_kpi(
        ws,
        "G5:I8",
        "LÍDER ESTRUCTURAL",
        lider["Departamento"],
        VERDE_OSCURO
    )

    bloque_kpi(
        ws,
        "J5:L8",
        "FORTALEZA CONSOLIDADA",
        consolidado,
        AZUL
    )

    bloque_kpi(
        ws,
        "M5:P8",
        "ALERTAS ROJAS",
        int(alertas_rojas),
        ROJO
    )

    # ========================================================
    # TOP 10 IRNA
    # ========================================================

    ws["A10"] = (
        "TOP 10 — IRNA ESTRUCTURAL"
    )

    ws["A10"].font = Font(
        size=14,
        bold=True,
        color=VERDE_OSCURO
    )

    ws.append([])

    headers = [
        "Departamento",
        "IRNA",
        "Ejecución",
        "Brecha",
        "Alerta"
    ]

    for i, h in enumerate(
        headers,
        start=1
    ):
        cell = ws.cell(
            11,
            i,
            h
        )

        cell.fill = PatternFill(
            "solid",
            fgColor=VERDE_OSCURO
        )

        cell.font = Font(
            bold=True,
            color=BLANCO
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    top10 = (
        ranking
        .sort_values(
            "IRNA_Estructural",
            ascending=False
        )
        .head(10)
    )

    fila = 12

    for _, row in top10.iterrows():

        ws.cell(
            fila,
            1,
            row["Departamento"]
        )

        ws.cell(
            fila,
            2,
            row["IRNA_Estructural"]
        )

        ws.cell(
            fila,
            3,
            row["IRNA_Ejecucion"]
        )

        ws.cell(
            fila,
            4,
            row["Brecha_IRNA"]
        )

        alerta = ws.cell(
            fila,
            5,
            row["Nivel_Alerta"]
        )

        alerta.fill = PatternFill(
            "solid",
            fgColor=color_alerta(
                row["Nivel_Alerta"]
            )
        )

        alerta.font = Font(
            bold=True,
            color=BLANCO
        )

        alerta.alignment = Alignment(
            horizontal="center"
        )

        fila += 1

    # ========================================================
    # GRÁFICO TOP 10
    # ========================================================

    chart = BarChart()

    chart.type = "bar"
    chart.style = 10
    chart.title = "Top 10 IRNA estructural"
    chart.x_axis.title = "IRNA"
    chart.y_axis.title = "Departamento"

    data = Reference(
        ws,
        min_col=2,
        min_row=11,
        max_row=21
    )

    cats = Reference(
        ws,
        min_col=1,
        min_row=12,
        max_row=21
    )

    chart.add_data(
        data,
        titles_from_data=True
    )

    chart.set_categories(
        cats
    )

    chart.height = 9
    chart.width = 16

    chart.legend = None

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    ws.add_chart(
        chart,
        "G10"
    )

    # ========================================================
    # BRECHAS
    # ========================================================

    ws["A24"] = (
        "TOP BRECHAS ENTRE POTENCIAL Y EJECUCIÓN"
    )

    ws["A24"].font = Font(
        size=14,
        bold=True,
        color=ROJO
    )

    for i, h in enumerate(
        [
            "Departamento",
            "IRNA Estructural",
            "Ejecución",
            "Brecha"
        ],
        start=1
    ):

        cell = ws.cell(
            25,
            i,
            h
        )

        cell.fill = PatternFill(
            "solid",
            fgColor=ROJO
        )

        cell.font = Font(
            bold=True,
            color=BLANCO
        )

    top_brechas = (
        ranking
        .sort_values(
            "Brecha_IRNA",
            ascending=False
        )
        .head(8)
    )

    fila = 26

    for _, row in top_brechas.iterrows():

        ws.cell(
            fila,
            1,
            row["Departamento"]
        )

        ws.cell(
            fila,
            2,
            row["IRNA_Estructural"]
        )

        ws.cell(
            fila,
            3,
            row["IRNA_Ejecucion"]
        )

        ws.cell(
            fila,
            4,
            row["Brecha_IRNA"]
        )

        fila += 1

    chart_brecha = BarChart()

    chart_brecha.type = "bar"
    chart_brecha.style = 10
    chart_brecha.title = (
        "Principales brechas territoriales"
    )

    chart_brecha.x_axis.title = (
        "Brecha IRNA"
    )

    data = Reference(
        ws,
        min_col=4,
        min_row=25,
        max_row=33
    )

    cats = Reference(
        ws,
        min_col=1,
        min_row=26,
        max_row=33
    )

    chart_brecha.add_data(
        data,
        titles_from_data=True
    )

    chart_brecha.set_categories(
        cats
    )

    chart_brecha.height = 8
    chart_brecha.width = 16
    chart_brecha.legend = None

    chart_brecha.dataLabels = (
        DataLabelList()
    )

    chart_brecha.dataLabels.showVal = True

    ws.add_chart(
        chart_brecha,
        "G25"
    )

    # ========================================================
    # SCATTER CORRECTO
    # ========================================================

    ws["A36"] = (
        "POTENCIAL ESTRUCTURAL VS EJECUCIÓN"
    )

    ws["A36"].font = Font(
        size=14,
        bold=True,
        color=AZUL
    )

    # Crear tabla auxiliar invisible más abajo
    fila_inicio = 70

    ws.cell(
        fila_inicio,
        1,
        "Departamento"
    )

    ws.cell(
        fila_inicio,
        2,
        "IRNA_Estructural"
    )

    ws.cell(
        fila_inicio,
        3,
        "IRNA_Ejecucion"
    )

    for i, row in ranking.iterrows():

        f = fila_inicio + 1 + i

        ws.cell(
            f,
            1,
            row["Departamento"]
        )

        ws.cell(
            f,
            2,
            row["IRNA_Estructural"]
        )

        ws.cell(
            f,
            3,
            row["IRNA_Ejecucion"]
        )

    scatter = ScatterChart()

    scatter.title = (
        "IRNA estructural vs ejecución"
    )

    scatter.x_axis.title = (
        "IRNA Estructural"
    )

    scatter.y_axis.title = (
        "IRNA Ejecución"
    )

    scatter.height = 11
    scatter.width = 18

    xvalues = Reference(
        ws,
        min_col=2,
        min_row=fila_inicio + 1,
        max_row=fila_inicio + len(ranking)
    )

    yvalues = Reference(
        ws,
        min_col=3,
        min_row=fila_inicio + 1,
        max_row=fila_inicio + len(ranking)
    )

    serie = Series(
        yvalues,
        xvalues,
        title="Territorios"
    )

    # solo marcadores, sin líneas
    serie.graphicalProperties.line.noFill = True

    serie.marker.symbol = "circle"
    serie.marker.size = 7

    scatter.series.append(
        serie
    )

    scatter.legend = None

    ws.add_chart(
        scatter,
        "A38"
    )

    # ========================================================
    # LECTURA ESTRATÉGICA
    # ========================================================

    ws.merge_cells(
        "J38:P52"
    )

    mejores_ejecucion = (
        ranking
        .sort_values(
            "IRNA_Ejecucion",
            ascending=False
        )
        .head(2)["Departamento"]
        .tolist()
    )

    lectura = (
        "LECTURA ESTRATÉGICA\n\n"
        f"• {lider['Departamento']} lidera la fortaleza estructural "
        f"con IRNA {lider['IRNA_Estructural']:.1f}, pero presenta una "
        f"brecha de {lider['Brecha_IRNA']:.1f} puntos.\n\n"
        f"• {consolidado} aparece como la principal fortaleza consolidada "
        f"del Radar.\n\n"
        f"• {mejores_ejecucion[0]} y {mejores_ejecucion[1]} presentan "
        f"los mayores niveles de ejecución del núcleo Radar.\n\n"
        f"• Existen {alertas_rojas} territorios en alerta roja que requieren "
        f"seguimiento prioritario de inversión y gestión."
    )

    ws["J38"] = lectura

    ws["J38"].fill = PatternFill(
        "solid",
        fgColor=GRIS_CLARO
    )

    ws["J38"].font = Font(
        size=11,
        color=GRIS_OSCURO
    )

    ws["J38"].alignment = Alignment(
        vertical="top",
        wrap_text=True
    )

    # ========================================================
    # NOTA METODOLÓGICA
    # ========================================================

    ws.merge_cells(
        "A56:P61"
    )

    ws["A56"] = (
        "NOTA METODOLÓGICA\n"
        "IRNA Estructural integra inversión, vocación territorial, masa crítica, "
        "diversidad y evidencia. IRNA Ejecución refleja el avance presupuestal. "
        "La Brecha IRNA mide la diferencia entre fortaleza estructural y ejecución. "
        "Los cuadrantes se interpretan respecto de las medianas nacionales "
        f"(IRNA estructural {mediana_estructural:.1f}; ejecución {mediana_ejecucion:.1f})."
    )

    ws["A56"].fill = PatternFill(
        "solid",
        fgColor=GRIS_CLARO
    )

    ws["A56"].font = Font(
        size=10,
        color=GRIS_OSCURO
    )

    ws["A56"].alignment = Alignment(
        vertical="top",
        wrap_text=True
    )

    # ========================================================
    # FORMATO GENERAL
    # ========================================================

    widths = {
        "A": 23,
        "B": 13,
        "C": 13,
        "D": 13,
        "E": 12,
        "F": 3,
        "G": 13,
        "H": 13,
        "I": 13,
        "J": 15,
        "K": 15,
        "L": 15,
        "M": 15,
        "N": 15,
        "O": 15,
        "P": 15
    }

    for col, width in widths.items():
        ws.column_dimensions[
            col
        ].width = width

    for row in range(
        1,
        62
    ):
        ws.row_dimensions[
            row
        ].height = 22

    ws.row_dimensions[
        1
    ].height = 32

    # ocultar tabla auxiliar scatter
    for row in range(
        fila_inicio,
        fila_inicio + len(ranking) + 2
    ):
        ws.row_dimensions[
            row
        ].hidden = True

    # ========================================================
    # FORMATO HOJAS AUXILIARES
    # ========================================================

    for nombre in [
        "Ranking",
        "Potencial_Ejecucion",
        "Alertas",
        "Oportunidades",
        "Indicadores"
    ]:

        hoja = wb[nombre]

        hoja.freeze_panes = "A2"

        hoja.auto_filter.ref = (
            hoja.dimensions
        )

        estilo_encabezado(
            hoja
        )

        ajustar_columnas(
            hoja
        )

        for row in hoja.iter_rows(
            min_row=2
        ):

            for cell in row:

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )

    # ========================================================
    # GUARDAR
    # ========================================================

    wb.save(
        SALIDA
    )

    print("\n" + "=" * 100)
    print("DASHBOARD PROFESIONAL GENERADO")
    print("=" * 100)

    print(SALIDA)

    print(
        "\n✓ DASHBOARD PROFESIONAL COMPLETADO"
    )

    print("FIN")


if __name__ == "__main__":
    main()