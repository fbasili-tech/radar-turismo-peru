from pathlib import Path
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList


# ============================================================
# RADAR DE TURISMO DE NATURALEZA Y AVENTURA DEL PERÚ
# ETAPA 16 - RADAR EJECUTIVO NACIONAL
# VERSIÓN CORREGIDA
# ============================================================

BASE_TABLERO = Path(
    "outputs/radar_turismo_tablero_base_2026.xlsx"
)

BASE_SEMAFORO = Path(
    "outputs/radar_semaforo_territorial_2026.xlsx"
)

SALIDA = Path(
    "outputs/radar_ejecutivo_nacional_2026.xlsx"
)


# ============================================================
# PALETA
# ============================================================

VERDE_OSCURO = "143D33"
VERDE = "2E7D5B"
VERDE_CLARO = "65B68A"

AZUL = "2F6B9A"
AZUL_OSCURO = "174A73"

ROJO = "D9534F"
NARANJA = "F39C45"
AMARILLO = "E6C84F"

GRIS = "C9CED3"
GRIS_CLARO = "F3F5F7"
GRIS_TEXTO = "404040"

BLANCO = "FFFFFF"


# ============================================================
# FUNCIONES DE ESTILO
# ============================================================

def fill(color):
    return PatternFill(
        "solid",
        fgColor=color
    )


def borde():

    lado = Side(
        style="thin",
        color="D9D9D9"
    )

    return Border(
        left=lado,
        right=lado,
        top=lado,
        bottom=lado
    )


def color_alerta(alerta):

    colores = {
        "ROJA": ROJO,
        "NARANJA": NARANJA,
        "AMARILLA": AMARILLO,
        "VERDE": VERDE
    }

    return colores.get(
        str(alerta).upper(),
        GRIS
    )


def color_estado(estado):

    colores = {
        "FORTALEZA CONSOLIDADA": VERDE_OSCURO,
        "BUENA EJECUCIÓN": VERDE_CLARO,
        "EN CONSOLIDACIÓN": AMARILLO,
        "REQUIERE ACELERACIÓN": NARANJA,
        "BRECHA CRÍTICA": ROJO,
        "SIN EVIDENCIA RADAR": GRIS
    }

    return colores.get(
        str(estado).upper(),
        GRIS
    )


def titulo_seccion(
    ws,
    fila,
    texto,
    color=VERDE_OSCURO
):

    ws.merge_cells(
        start_row=fila,
        start_column=1,
        end_row=fila,
        end_column=16
    )

    celda = ws.cell(
        fila,
        1,
        texto
    )

    celda.font = Font(
        size=14,
        bold=True,
        color=color
    )

    celda.alignment = Alignment(
        vertical="center"
    )

    ws.row_dimensions[
        fila
    ].height = 26


def kpi(
    ws,
    rango,
    titulo,
    valor,
    color
):

    ws.merge_cells(
        rango
    )

    celda = ws[
        rango.split(":")[0]
    ]

    celda.value = (
        f"{titulo}\n{valor}"
    )

    celda.fill = fill(
        color
    )

    celda.font = Font(
        size=13,
        bold=True,
        color=BLANCO
    )

    celda.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )


# ============================================================
# PROCESO PRINCIPAL
# ============================================================

def main():

    print("=" * 100)
    print(
        "RADAR DE TURISMO DE NATURALEZA "
        "Y AVENTURA DEL PERÚ"
    )
    print(
        "16 - RADAR EJECUTIVO NACIONAL"
    )
    print("=" * 100)

    # ========================================================
    # CARGAR DATOS
    # ========================================================

    ranking = pd.read_excel(
        BASE_TABLERO,
        sheet_name="Ranking_Nacional"
    )

    alertas = pd.read_excel(
        BASE_TABLERO,
        sheet_name="Alertas_Territoriales"
    )

    oportunidades = pd.read_excel(
        BASE_TABLERO,
        sheet_name="Oportunidades"
    )

    semaforo = pd.read_excel(
        BASE_SEMAFORO,
        sheet_name="Mapa_Semaforo"
    )

    print(
        f"\nTerritorios con evidencia Radar: "
        f"{len(ranking)}"
    )

    print(
        f"Territorios del mapa nacional: "
        f"{len(semaforo)}"
    )

    # ========================================================
    # ORDENAR RANKING
    # ========================================================

    ranking = (
        ranking
        .sort_values(
            "IRNA_Estructural",
            ascending=False
        )
        .reset_index(
            drop=True
        )
    )

    # ========================================================
    # RECONSTRUIR DEVENGADO
    # ========================================================

    ranking[
        "Devengado_Radar_Calculado"
    ] = (
        ranking["PIM_Radar"]
        * ranking["IRNA_Ejecucion"]
        / 100
    )

    # ========================================================
    # INDICADORES NACIONALES
    # ========================================================

    pim_radar = ranking[
        "PIM_Radar"
    ].sum()

    devengado_radar = ranking[
        "Devengado_Radar_Calculado"
    ].sum()

    avance_nacional = (
        devengado_radar
        / pim_radar
        * 100
        if pim_radar > 0
        else 0
    )

    territorios = len(
        ranking
    )

    lider = ranking.iloc[
        0
    ]

    rojas = (
        ranking[
            "Nivel_Alerta"
        ]
        == "ROJA"
    ).sum()

    naranjas = (
        ranking[
            "Nivel_Alerta"
        ]
        == "NARANJA"
    ).sum()

    sin_evidencia = (
        semaforo[
            "Estado_Semaforo"
        ]
        == "SIN EVIDENCIA RADAR"
    ).sum()

    mediana_irna = ranking[
        "IRNA_Estructural"
    ].median()

    mediana_ejecucion = ranking[
        "IRNA_Ejecucion"
    ].median()

    fortalezas = ranking[
        ranking[
            "Estado_Territorial"
        ]
        == "FORTALEZA CONSOLIDADA"
    ]

    if len(
        fortalezas
    ) > 0:

        fortaleza = fortalezas.iloc[
            0
        ][
            "Departamento"
        ]

    else:

        fortaleza = "—"

    # ========================================================
    # MAYOR BRECHA / MEJOR EJECUCIÓN
    # ========================================================

    mayor_brecha = (
        ranking
        .sort_values(
            "Brecha_IRNA",
            ascending=False
        )
        .iloc[0]
    )

    mejor_ejecucion = (
        ranking
        .sort_values(
            "IRNA_Ejecucion",
            ascending=False
        )
        .iloc[0]
    )

    # ========================================================
    # CREAR LIBRO
    # ========================================================

    wb = Workbook()

    ws = wb.active

    ws.title = (
        "Radar_Ejecutivo"
    )

    ws.sheet_view.showGridLines = False

    # ========================================================
    # TÍTULO
    # ========================================================

    ws.merge_cells(
        "A1:P2"
    )

    ws["A1"] = (
        "RADAR TERRITORIAL DE INVERSIÓN PÚBLICA "
        "EN TURISMO DE NATURALEZA Y AVENTURA — PERÚ 2026"
    )

    ws["A1"].fill = fill(
        VERDE_OSCURO
    )

    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=BLANCO
    )

    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    ws.merge_cells(
        "A3:P3"
    )

    ws["A3"] = (
        "IRNA | Fortaleza estructural, "
        "ejecución presupuestal y brechas de gestión"
    )

    ws["A3"].font = Font(
        size=11,
        italic=True,
        color=GRIS_TEXTO
    )

    ws["A3"].alignment = Alignment(
        horizontal="center"
    )

    # ========================================================
    # KPIs PRINCIPALES
    # ========================================================

    kpi(
        ws,
        "A5:C8",
        "TERRITORIOS CON EVIDENCIA",
        territorios,
        AZUL
    )

    kpi(
        ws,
        "D5:F8",
        "PIM NÚCLEO RADAR",
        f"S/ {pim_radar / 1_000_000:.1f} M",
        VERDE
    )

    kpi(
        ws,
        "G5:I8",
        "AVANCE PRESUPUESTAL",
        f"{avance_nacional:.1f}%",
        AZUL_OSCURO
    )

    kpi(
        ws,
        "J5:L8",
        "LÍDER ESTRUCTURAL",
        lider["Departamento"],
        VERDE_OSCURO
    )

    kpi(
        ws,
        "M5:P8",
        "ALERTAS ROJAS",
        int(rojas),
        ROJO
    )

    # ========================================================
    # SEGUNDA LÍNEA DE KPIs
    # ========================================================

    kpi(
        ws,
        "A10:C12",
        "FORTALEZA CONSOLIDADA",
        fortaleza,
        VERDE
    )

    kpi(
        ws,
        "D10:F12",
        "MEDIANA IRNA",
        f"{mediana_irna:.1f}",
        AZUL
    )

    kpi(
        ws,
        "G10:I12",
        "MEDIANA EJECUCIÓN",
        f"{mediana_ejecucion:.1f}%",
        AZUL
    )

    kpi(
        ws,
        "J10:L12",
        "ALERTAS NARANJAS",
        int(naranjas),
        NARANJA
    )

    kpi(
        ws,
        "M10:P12",
        "SIN EVIDENCIA RADAR",
        int(sin_evidencia),
        GRIS
    )

    # ========================================================
    # LECTURA NACIONAL
    # ========================================================

    titulo_seccion(
        ws,
        14,
        "LECTURA NACIONAL DEL RADAR"
    )

    ws.merge_cells(
        "A15:P20"
    )

    ws["A15"] = (
        f"El núcleo Radar identifica S/ "
        f"{pim_radar / 1_000_000:.1f} millones "
        f"de inversión pública vinculada al turismo de "
        f"naturaleza y aventura en {territorios} territorios. "
        f"El avance presupuestal agregado alcanza "
        f"{avance_nacional:.1f}%.\n\n"

        f"{lider['Departamento']} encabeza la fortaleza "
        f"estructural con un IRNA de "
        f"{lider['IRNA_Estructural']:.1f}. "
        f"{fortaleza} aparece como fortaleza consolidada. "
        f"La mayor brecha entre fortaleza estructural y "
        f"ejecución corresponde a "
        f"{mayor_brecha['Departamento']} "
        f"({mayor_brecha['Brecha_IRNA']:.1f} puntos), "
        f"mientras que {mejor_ejecucion['Departamento']} "
        f"registra la mayor ejecución del núcleo "
        f"({mejor_ejecucion['IRNA_Ejecucion']:.1f}%).\n\n"

        f"El Radar registra {int(rojas)} alertas rojas y "
        f"{int(naranjas)} alertas naranjas. "
        f"{int(sin_evidencia)} territorios aparecen como "
        f"sin evidencia suficiente dentro del núcleo "
        f"de inversión analizado."
    )

    ws["A15"].fill = fill(
        GRIS_CLARO
    )

    ws["A15"].font = Font(
        size=11,
        color=GRIS_TEXTO
    )

    ws["A15"].alignment = Alignment(
        vertical="top",
        wrap_text=True
    )

    # ========================================================
    # TOP 10 IRNA
    # ========================================================

    titulo_seccion(
        ws,
        22,
        "TOP 10 — FORTALEZA ESTRUCTURAL IRNA"
    )

    headers = [
        "Ranking",
        "Departamento",
        "IRNA",
        "Ejecución",
        "Brecha",
        "Alerta"
    ]

    for col, texto in enumerate(
        headers,
        start=1
    ):

        celda = ws.cell(
            23,
            col,
            texto
        )

        celda.fill = fill(
            VERDE_OSCURO
        )

        celda.font = Font(
            bold=True,
            color=BLANCO
        )

        celda.alignment = Alignment(
            horizontal="center"
        )

    top10 = ranking.head(
        10
    )

    for pos, (_, row) in enumerate(
        top10.iterrows(),
        start=1
    ):

        fila = 23 + pos

        ws.cell(
            fila,
            1,
            pos
        )

        ws.cell(
            fila,
            2,
            row["Departamento"]
        )

        ws.cell(
            fila,
            3,
            row["IRNA_Estructural"]
        )

        ws.cell(
            fila,
            4,
            row["IRNA_Ejecucion"]
        )

        ws.cell(
            fila,
            5,
            row["Brecha_IRNA"]
        )

        alerta = ws.cell(
            fila,
            6,
            row["Nivel_Alerta"]
        )

        alerta.fill = fill(
            color_alerta(
                row["Nivel_Alerta"]
            )
        )

        alerta.font = Font(
            bold=True,
            color=BLANCO
        )

        alerta.alignment = Alignment(
            horizontal="center"
        )

    # ========================================================
    # GRÁFICO TOP 10
    # ========================================================

    chart = BarChart()

    chart.type = "bar"

    chart.title = (
        "Top 10 IRNA estructural"
    )

    chart.x_axis.title = (
        "IRNA"
    )

    chart.y_axis.title = None

    datos = Reference(
        ws,
        min_col=3,
        min_row=24,
        max_row=33
    )

    categorias = Reference(
        ws,
        min_col=2,
        min_row=24,
        max_row=33
    )

    chart.add_data(
        datos,
        titles_from_data=False,
        from_rows=False
    )

    chart.set_categories(
        categorias
    )

    chart.legend = None

    chart.height = 9
    chart.width = 17

    chart.dataLabels = (
        DataLabelList()
    )

    chart.dataLabels.showVal = True
    chart.dataLabels.showLegendKey = False
    chart.dataLabels.showCatName = False
    chart.dataLabels.showSerName = False

    chart.series[
        0
    ].graphicalProperties.solidFill = VERDE

    chart.series[
        0
    ].graphicalProperties.line.solidFill = VERDE

    ws.add_chart(
        chart,
        "H23"
    )

    # ========================================================
    # BRECHAS PRIORITARIAS
    # ========================================================

    titulo_seccion(
        ws,
        36,
        "BRECHAS TERRITORIALES PRIORITARIAS",
        ROJO
    )

    headers_b = [
        "Departamento",
        "IRNA",
        "Ejecución",
        "Brecha",
        "Estado"
    ]

    for col, texto in enumerate(
        headers_b,
        start=1
    ):

        celda = ws.cell(
            37,
            col,
            texto
        )

        celda.fill = fill(
            ROJO
        )

        celda.font = Font(
            bold=True,
            color=BLANCO
        )

    brechas = (
        ranking
        .sort_values(
            "Brecha_IRNA",
            ascending=False
        )
        .head(8)
    )

    for i, (_, row) in enumerate(
        brechas.iterrows(),
        start=38
    ):

        ws.cell(
            i,
            1,
            row["Departamento"]
        )

        ws.cell(
            i,
            2,
            row["IRNA_Estructural"]
        )

        ws.cell(
            i,
            3,
            row["IRNA_Ejecucion"]
        )

        ws.cell(
            i,
            4,
            row["Brecha_IRNA"]
        )

        ws.cell(
            i,
            5,
            row["Estado_Territorial"]
        )

    # ========================================================
    # GRÁFICO BRECHAS
    # ========================================================

    chart_b = BarChart()

    chart_b.type = "bar"

    chart_b.title = (
        "Mayores brechas IRNA"
    )

    chart_b.x_axis.title = (
        "Brecha"
    )

    datos_b = Reference(
        ws,
        min_col=4,
        min_row=38,
        max_row=45
    )

    categorias_b = Reference(
        ws,
        min_col=1,
        min_row=38,
        max_row=45
    )

    chart_b.add_data(
        datos_b,
        titles_from_data=False,
        from_rows=False
    )

    chart_b.set_categories(
        categorias_b
    )

    chart_b.legend = None

    chart_b.height = 8
    chart_b.width = 17

    chart_b.dataLabels = (
        DataLabelList()
    )

    chart_b.dataLabels.showVal = True
    chart_b.dataLabels.showLegendKey = False
    chart_b.dataLabels.showCatName = False
    chart_b.dataLabels.showSerName = False

    chart_b.series[
        0
    ].graphicalProperties.solidFill = ROJO

    chart_b.series[
        0
    ].graphicalProperties.line.solidFill = ROJO

    ws.add_chart(
        chart_b,
        "H37"
    )

    # ========================================================
    # SEMÁFORO NACIONAL
    # ========================================================

    titulo_seccion(
        ws,
        49,
        "SEMÁFORO TERRITORIAL NACIONAL"
    )

    resumen_semaforo = (
        semaforo
        .groupby(
            "Estado_Semaforo",
            as_index=False
        )
        .agg(
            Territorios=(
                "Departamento_Mapa",
                "count"
            ),
            PIM_Radar=(
                "PIM_Radar",
                "sum"
            )
        )
    )

    headers_s = [
        "Estado territorial",
        "Territorios",
        "PIM Radar"
    ]

    for col, texto in enumerate(
        headers_s,
        start=1
    ):

        celda = ws.cell(
            50,
            col,
            texto
        )

        celda.fill = fill(
            VERDE_OSCURO
        )

        celda.font = Font(
            bold=True,
            color=BLANCO
        )

    for i, (_, row) in enumerate(
        resumen_semaforo.iterrows(),
        start=51
    ):

        estado = row[
            "Estado_Semaforo"
        ]

        celda = ws.cell(
            i,
            1,
            estado
        )

        celda.fill = fill(
            color_estado(
                estado
            )
        )

        celda.font = Font(
            bold=True,
            color=BLANCO
        )

        ws.cell(
            i,
            2,
            row["Territorios"]
        )

        ws.cell(
            i,
            3,
            row["PIM_Radar"]
        )

        ws.cell(
            i,
            3
        ).number_format = (
            '"S/ " #,##0'
        )

    # ========================================================
    # PRIORIDADES DE GESTIÓN
    # ========================================================

    titulo_seccion(
        ws,
        60,
        "PRIORIDADES DE GESTIÓN 2026"
    )

    top_brecha = (
        ranking
        .sort_values(
            "Brecha_IRNA",
            ascending=False
        )
        .head(5)
        ["Departamento"]
        .tolist()
    )

    top_oportunidad = (
        ranking[
            (
                ranking[
                    "IRNA_Estructural"
                ]
                >= mediana_irna
            )
            &
            (
                ranking[
                    "IRNA_Ejecucion"
                ]
                >= mediana_ejecucion
            )
        ]
        .sort_values(
            "IRNA_Estructural",
            ascending=False
        )
        .head(5)
        ["Departamento"]
        .tolist()
    )

    ws.merge_cells(
        "A61:P69"
    )

    ws["A61"] = (
        "1. DESTRABE Y ACELERACIÓN\n"
        f"Prioridad territorial: "
        f"{', '.join(top_brecha)}.\n\n"

        "2. CONSOLIDACIÓN DE FORTALEZAS\n"
        f"Territorios con combinación favorable de "
        f"estructura y ejecución: "
        f"{', '.join(top_oportunidad)}.\n\n"

        "3. GENERACIÓN DE EVIDENCIA\n"
        "Huancavelica y Pasco deben diferenciarse "
        "explícitamente como territorios sin evidencia "
        "suficiente en el núcleo Radar 2026, no como "
        "territorios de bajo potencial.\n\n"

        "4. SIGUIENTE CAPA DEL RADAR\n"
        "Integrar progresivamente demanda turística, "
        "áreas naturales protegidas, conectividad, "
        "oferta empresarial, seguridad, empleo y "
        "mercado para evolucionar desde un radar de "
        "inversión pública hacia un sistema nacional "
        "de inteligencia del turismo de naturaleza "
        "y aventura."
    )

    ws["A61"].fill = fill(
        GRIS_CLARO
    )

    ws["A61"].font = Font(
        size=11,
        color=GRIS_TEXTO
    )

    ws["A61"].alignment = Alignment(
        vertical="top",
        wrap_text=True
    )

    # ========================================================
    # NOTA METODOLÓGICA
    # ========================================================

    ws.merge_cells(
        "A72:P77"
    )

    ws["A72"] = (
        "ALCANCE METODOLÓGICO\n"
        "El Radar 2026 analiza inversión pública identificada "
        "como directa o complementariamente vinculada al "
        "turismo de naturaleza y aventura. El IRNA no debe "
        "interpretarse como un índice general de competitividad "
        "turística regional. Su función es comparar fortaleza "
        "estructural de la cartera observada, ejecución "
        "presupuestal, evidencia y brechas de gestión dentro "
        "del universo analizado."
    )

    ws["A72"].fill = fill(
        "E8EEF2"
    )

    ws["A72"].font = Font(
        size=10,
        italic=True,
        color=GRIS_TEXTO
    )

    ws["A72"].alignment = Alignment(
        vertical="top",
        wrap_text=True
    )

    # ========================================================
    # FORMATOS GENERALES
    # ========================================================

    anchos = {
        "A": 24,
        "B": 18,
        "C": 15,
        "D": 15,
        "E": 30,
        "F": 13,
        "G": 3,
        "H": 15,
        "I": 15,
        "J": 15,
        "K": 15,
        "L": 15,
        "M": 15,
        "N": 15,
        "O": 15,
        "P": 15
    }

    for col, ancho in anchos.items():

        ws.column_dimensions[
            col
        ].width = ancho

    for fila in range(
        1,
        78
    ):

        ws.row_dimensions[
            fila
        ].height = 22

    ws.row_dimensions[
        1
    ].height = 34

    # Bordes top 10
    for fila in range(
        24,
        34
    ):

        for col in range(
            1,
            7
        ):

            ws.cell(
                fila,
                col
            ).border = borde()

    # Bordes brechas
    for fila in range(
        38,
        46
    ):

        for col in range(
            1,
            6
        ):

            ws.cell(
                fila,
                col
            ).border = borde()

    # ========================================================
    # HOJA RANKING NACIONAL
    # ========================================================

    ws_rank = wb.create_sheet(
        "Ranking_Nacional"
    )

    for c_idx, columna in enumerate(
        ranking.columns,
        start=1
    ):

        ws_rank.cell(
            1,
            c_idx,
            columna
        )

    for r_idx, row in enumerate(
        ranking.itertuples(
            index=False,
            name=None
        ),
        start=2
    ):

        for c_idx, valor in enumerate(
            row,
            start=1
        ):

            ws_rank.cell(
                r_idx,
                c_idx,
                valor
            )

    # ========================================================
    # HOJA SEMÁFORO
    # ========================================================

    ws_sem = wb.create_sheet(
        "Semaforo_Territorial"
    )

    for c_idx, columna in enumerate(
        semaforo.columns,
        start=1
    ):

        ws_sem.cell(
            1,
            c_idx,
            columna
        )

    for r_idx, row in enumerate(
        semaforo.itertuples(
            index=False,
            name=None
        ),
        start=2
    ):

        for c_idx, valor in enumerate(
            row,
            start=1
        ):

            ws_sem.cell(
                r_idx,
                c_idx,
                valor
            )

    # ========================================================
    # GUARDAR
    # ========================================================

    wb.save(
        SALIDA
    )

    # ========================================================
    # CONTROL FINAL
    # ========================================================

    print(
        "\n" + "=" * 100
    )

    print(
        "RADAR EJECUTIVO NACIONAL GENERADO"
    )

    print(
        "=" * 100
    )

    print(
        f"\nPIM núcleo Radar : "
        f"S/ {pim_radar:,.0f}"
    )

    print(
        f"Devengado calc.  : "
        f"S/ {devengado_radar:,.0f}"
    )

    print(
        f"Avance nacional  : "
        f"{avance_nacional:.1f}%"
    )

    print(
        f"Territorios      : "
        f"{territorios}"
    )

    print(
        f"Alertas rojas    : "
        f"{int(rojas)}"
    )

    print(
        f"Alertas naranjas : "
        f"{int(naranjas)}"
    )

    print(
        f"Sin evidencia    : "
        f"{int(sin_evidencia)}"
    )

    print(
        f"\nARCHIVO GENERADO:\n"
        f"{SALIDA}"
    )

    print(
        "\n✓ RADAR EJECUTIVO NACIONAL COMPLETADO"
    )

    print(
        "FIN"
    )


if __name__ == "__main__":
    main()