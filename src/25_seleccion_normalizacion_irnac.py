from pathlib import Path

import numpy as np
import pandas as pd


# ============================================================
# RADAR TURISMO NATURALEZA Y AVENTURA - PERÚ
#
# 25 - SELECCIÓN Y NORMALIZACIÓN DE INDICADORES IRNA-C
#
# OBJETIVO
# ------------------------------------------------------------
# Seleccionar indicadores no redundantes y normalizarlos
# a una escala común 0-100.
#
# Esta etapa:
# - NO calcula todavía el IRNA-C final.
# - NO asigna pesos definitivos.
# - NO genera ranking oficial.
# ============================================================


# ============================================================
# ARCHIVOS
# ============================================================

ENTRADA = Path(
    "outputs/radar_fase2_matriz_maestra_2026.xlsx"
)

SALIDA = Path(
    "outputs/radar_fase2_indicadores_normalizados_irnac_2026.xlsx"
)

SALIDA.parent.mkdir(
    parents=True,
    exist_ok=True
)


# ============================================================
# HOJA DE ENTRADA
# ============================================================

HOJA = "02_Matriz_Analitica"


# ============================================================
# PARÁMETROS DE NORMALIZACIÓN
# ============================================================

PERCENTIL_INFERIOR = 0.05
PERCENTIL_SUPERIOR = 0.95


# ============================================================
# ARQUITECTURA PROPUESTA DEL IRNA-C
# ============================================================

INDICADORES = [

    # --------------------------------------------------------
    # 1. DEMANDA
    # --------------------------------------------------------

    {
        "Variable":
            "Demanda_Visitantes_Sitios",

        "Dimension":
            "DEMANDA",

        "Indicador":
            "Volumen de demanda turística",

        "Sentido":
            "POSITIVO",

        "Decision":
            "INCLUIR",

        "Justificacion":
            (
                "Mide intensidad de visita a atractivos "
                "y sitios turísticos."
            )
    },

    {
        "Variable":
            "Demanda_Participacion_Extranjera_Porcentaje",

        "Dimension":
            "DEMANDA",

        "Indicador":
            "Internacionalización de la demanda",

        "Sentido":
            "POSITIVO",

        "Decision":
            "INCLUIR",

        "Justificacion":
            (
                "Mide capacidad de atraer visitantes "
                "extranjeros."
            )
    },


    # --------------------------------------------------------
    # 2. DESEMPEÑO TURÍSTICO
    # --------------------------------------------------------

    {
        "Variable":
            "Hosp_Arribos",

        "Dimension":
            "DESEMPENO_TURISTICO",

        "Indicador":
            "Arribos a establecimientos de hospedaje",

        "Sentido":
            "POSITIVO",

        "Decision":
            "INCLUIR",

        "Justificacion":
            (
                "Representa volumen territorial de demanda "
                "alojada. Se elige frente a pernoctaciones "
                "para evitar doble conteo."
            )
    },

    {
        "Variable":
            "Hosp_Pernoctaciones",

        "Dimension":
            "DESEMPENO_TURISTICO",

        "Indicador":
            "Pernoctaciones",

        "Sentido":
            "POSITIVO",

        "Decision":
            "CONTROL",

        "Justificacion":
            (
                "Alta correlación con arribos; se conserva "
                "como indicador de contraste."
            )
    },

    {
        "Variable":
            "Hosp_Permanencia_Promedio",

        "Dimension":
            "DESEMPENO_TURISTICO",

        "Indicador":
            "Permanencia promedio",

        "Sentido":
            "POSITIVO",

        "Decision":
            "INCLUIR",

        "Justificacion":
            (
                "Una mayor permanencia implica mayor "
                "profundidad de consumo turístico."
            )
    },

    {
        "Variable":
            "Hosp_TNOH_Promedio",

        "Dimension":
            "DESEMPENO_TURISTICO",

        "Indicador":
            "Ocupabilidad hotelera",

        "Sentido":
            "POSITIVO",

        "Decision":
            "INCLUIR",

        "Justificacion":
            (
                "Mide utilización efectiva de la capacidad "
                "de alojamiento."
            )
    },


    # --------------------------------------------------------
    # 3. CONECTIVIDAD
    # --------------------------------------------------------

    {
        "Variable":
            "Aereo_Pasajeros_Total",

        "Dimension":
            "CONECTIVIDAD",

        "Indicador":
            "Movimiento aéreo total",

        "Sentido":
            "POSITIVO",

        "Decision":
            "INCLUIR",

        "Justificacion":
            (
                "Mide escala real de la conectividad aérea "
                "territorial."
            )
    },

    {
        "Variable":
            "Aereo_Aeropuertos_Con_Registro",

        "Dimension":
            "CONECTIVIDAD",

        "Indicador":
            "Cobertura aeroportuaria",

        "Sentido":
            "POSITIVO",

        "Decision":
            "INCLUIR",

        "Justificacion":
            (
                "Mide presencia de infraestructura aérea "
                "con movimiento registrado."
            )
    },

    {
        "Variable":
            "Aereo_Participacion_Internacional_Porcentaje",

        "Dimension":
            "CONECTIVIDAD",

        "Indicador":
            "Internacionalización aérea",

        "Sentido":
            "POSITIVO",

        "Decision":
            "INCLUIR",

        "Justificacion":
            (
                "Se usa participación internacional en lugar "
                "de pasajeros internacionales absolutos para "
                "reducir redundancia."
            )
    },

    {
        "Variable":
            "Aereo_Pasajeros_Internacionales",

        "Dimension":
            "CONECTIVIDAD",

        "Indicador":
            "Pasajeros internacionales",

        "Sentido":
            "POSITIVO",

        "Decision":
            "CONTROL",

        "Justificacion":
            (
                "Correlación casi perfecta con participación "
                "internacional; no entra al índice."
            )
    },


    # --------------------------------------------------------
    # 4. OFERTA FORMAL
    # --------------------------------------------------------

    {
        "Variable":
            "Oferta_Hospedajes_Formales",

        "Dimension":
            "OFERTA_FORMAL",

        "Indicador":
            "Hospedajes formales",

        "Sentido":
            "POSITIVO",

        "Decision":
            "INCLUIR",

        "Justificacion":
            (
                "Representa capacidad formal de alojamiento."
            )
    },

    {
        "Variable":
            "Oferta_Agencias_Formales",

        "Dimension":
            "OFERTA_FORMAL",

        "Indicador":
            "Agencias de viajes formales",

        "Sentido":
            "POSITIVO",

        "Decision":
            "INCLUIR",

        "Justificacion":
            (
                "Representa capacidad de intermediación "
                "y comercialización turística."
            )
    },

    {
        "Variable":
            "Oferta_Restaurantes_Calificados",

        "Dimension":
            "OFERTA_FORMAL",

        "Indicador":
            "Restaurantes calificados",

        "Sentido":
            "POSITIVO",

        "Decision":
            "INCLUIR",

        "Justificacion":
            (
                "Representa servicios complementarios "
                "formales del destino."
            )
    },

    {
        "Variable":
            "Oferta_Prestadores_Formales_Total",

        "Dimension":
            "OFERTA_FORMAL",

        "Indicador":
            "Prestadores formales totales",

        "Sentido":
            "POSITIVO",

        "Decision":
            "CONTROL",

        "Justificacion":
            (
                "Es suma de componentes y por tanto "
                "redundante para el índice."
            )
    },


    # --------------------------------------------------------
    # 5. CAPITAL NATURAL
    # --------------------------------------------------------

    {
        "Variable":
            "Capital_ANP_Nacional",

        "Dimension":
            "CAPITAL_NATURAL",

        "Indicador":
            "ANP nacionales",

        "Sentido":
            "POSITIVO",

        "Decision":
            "INCLUIR",

        "Justificacion":
            (
                "Mide presencia de áreas protegidas "
                "de administración nacional."
            )
    },

    {
        "Variable":
            "Capital_ACR_Total",

        "Dimension":
            "CAPITAL_NATURAL",

        "Indicador":
            "Áreas de Conservación Regional",

        "Sentido":
            "POSITIVO",

        "Decision":
            "INCLUIR",

        "Justificacion":
            (
                "Captura capital natural gestionado "
                "a escala regional."
            )
    },

    {
        "Variable":
            "Capital_Diversidad_Categorias",

        "Dimension":
            "CAPITAL_NATURAL",

        "Indicador":
            "Diversidad de categorías de conservación",

        "Sentido":
            "POSITIVO",

        "Decision":
            "INCLUIR",

        "Justificacion":
            (
                "Evita medir únicamente cantidad de ANP "
                "y captura diversidad del capital natural."
            )
    },

    {
        "Variable":
            "Capital_Superficie_Atribuida_Referencial_Ha",

        "Dimension":
            "CAPITAL_NATURAL",

        "Indicador":
            "Superficie protegida referencial",

        "Sentido":
            "POSITIVO",

        "Decision":
            "INCLUIR",

        "Justificacion":
            (
                "Mide escala territorial del capital natural. "
                "La superficie sigue siendo referencial "
                "hasta contar con GIS exacto."
            )
    },

    {
        "Variable":
            "Capital_ANP_Total",

        "Dimension":
            "CAPITAL_NATURAL",

        "Indicador":
            "ANP totales",

        "Sentido":
            "POSITIVO",

        "Decision":
            "CONTROL",

        "Justificacion":
            (
                "Alta correlación con ACP; se conserva "
                "como control descriptivo."
            )
    },

    {
        "Variable":
            "Capital_ACP_Total",

        "Dimension":
            "CAPITAL_NATURAL",

        "Indicador":
            "Áreas de Conservación Privada",

        "Sentido":
            "POSITIVO",

        "Decision":
            "CONTROL",

        "Justificacion":
            (
                "Se conserva para análisis complementario "
                "por su fuerte correlación con ANP total."
            )
    },

    {
        "Variable":
            "Capital_Zonas_Reservadas",

        "Dimension":
            "CAPITAL_NATURAL",

        "Indicador":
            "Zonas Reservadas",

        "Sentido":
            "POSITIVO",

        "Decision":
            "CONTROL",

        "Justificacion":
            (
                "Alta concentración de ceros; útil para "
                "diagnóstico, no como indicador principal."
            )
    }
]


# ============================================================
# CARGA
# ============================================================

def cargar_datos():

    if not ENTRADA.exists():

        raise FileNotFoundError(
            f"No existe:\n{ENTRADA}"
        )

    xls = pd.ExcelFile(
        ENTRADA
    )

    if HOJA not in xls.sheet_names:

        raise ValueError(
            f"No existe la hoja {HOJA}"
        )

    df = pd.read_excel(
        ENTRADA,
        sheet_name=HOJA
    )

    if "Departamento" not in df.columns:

        raise ValueError(
            "No existe la columna Departamento."
        )

    print(
        f"Territorios cargados : {len(df)}"
    )

    print(
        f"Variables disponibles: "
        f"{len(df.columns) - 1}"
    )

    return df


# ============================================================
# VALIDAR INDICADORES
# ============================================================

def validar_indicadores(
    df,
    catalogo
):

    faltantes = []

    for _, fila in catalogo.iterrows():

        variable = fila[
            "Variable"
        ]

        if variable not in df.columns:

            faltantes.append(
                variable
            )

    if faltantes:

        print(
            "\n⚠ VARIABLES NO ENCONTRADAS:"
        )

        for variable in faltantes:

            print(
                f" - {variable}"
            )

    return faltantes


# ============================================================
# WINSORIZACIÓN
# ============================================================

def winsorizar(
    serie,
    p_inferior=PERCENTIL_INFERIOR,
    p_superior=PERCENTIL_SUPERIOR
):

    s = pd.to_numeric(
        serie,
        errors="coerce"
    )

    validos = s.dropna()

    if validos.empty:

        return (
            s,
            np.nan,
            np.nan
        )

    inferior = validos.quantile(
        p_inferior
    )

    superior = validos.quantile(
        p_superior
    )

    resultado = s.clip(
        lower=inferior,
        upper=superior
    )

    return (
        resultado,
        inferior,
        superior
    )


# ============================================================
# NORMALIZACIÓN MIN-MAX
# ============================================================

def normalizar_0_100(
    serie,
    sentido
):

    s = pd.to_numeric(
        serie,
        errors="coerce"
    )

    minimo = s.min()
    maximo = s.max()

    if (
        pd.isna(minimo)
        or pd.isna(maximo)
    ):

        return pd.Series(
            np.nan,
            index=s.index
        )

    if maximo == minimo:

        return pd.Series(
            50.0,
            index=s.index
        )

    normalizada = (
        (s - minimo)
        /
        (maximo - minimo)
        *
        100
    )

    if sentido == "NEGATIVO":

        normalizada = (
            100 - normalizada
        )

    return normalizada


# ============================================================
# PROCESAR INDICADORES
# ============================================================

def construir_normalizados(
    df,
    catalogo
):

    resultado = pd.DataFrame(
        {
            "Departamento":
                df[
                    "Departamento"
                ]
        }
    )

    controles = []

    for _, meta in catalogo.iterrows():

        variable = meta[
            "Variable"
        ]

        if variable not in df.columns:

            continue

        original = pd.to_numeric(
            df[
                variable
            ],
            errors="coerce"
        )

        (
            robusta,
            p05,
            p95
        ) = winsorizar(
            original
        )

        normalizada = normalizar_0_100(
            robusta,
            meta[
                "Sentido"
            ]
        )

        # ----------------------------------------------------
        # Guardar variable normalizada
        # ----------------------------------------------------

        nombre_normalizado = (
            "N_"
            + variable
        )

        resultado[
            nombre_normalizado
        ] = normalizada

        # ----------------------------------------------------
        # Controles
        # ----------------------------------------------------

        modificados = int(
            (
                original.notna()
                &
                robusta.notna()
                &
                (
                    abs(
                        original - robusta
                    ) > 1e-12
                )
            ).sum()
        )

        controles.append(
            {
                "Variable":
                    variable,

                "Dimension":
                    meta[
                        "Dimension"
                    ],

                "Indicador":
                    meta[
                        "Indicador"
                    ],

                "Decision":
                    meta[
                        "Decision"
                    ],

                "Sentido":
                    meta[
                        "Sentido"
                    ],

                "P05":
                    p05,

                "P95":
                    p95,

                "Min_Original":
                    original.min(),

                "Max_Original":
                    original.max(),

                "Min_Normalizado":
                    normalizada.min(),

                "Max_Normalizado":
                    normalizada.max(),

                "Territorios_Modificados_Winsor":
                    modificados,

                "Faltantes":
                    int(
                        original.isna().sum()
                    )
            }
        )

    return (
        resultado,
        pd.DataFrame(
            controles
        )
    )


# ============================================================
# MATRIZ SOLO INDICADORES INCLUIDOS
# ============================================================

def construir_matriz_incluidos(
    normalizados,
    catalogo
):

    columnas = [
        "Departamento"
    ]

    incluidos = catalogo[
        catalogo[
            "Decision"
        ] == "INCLUIR"
    ]

    for variable in incluidos[
        "Variable"
    ]:

        columna = (
            "N_"
            + variable
        )

        if columna in normalizados.columns:

            columnas.append(
                columna
            )

    return normalizados[
        columnas
    ].copy()


# ============================================================
# RESUMEN POR DIMENSIÓN
# ============================================================

def construir_resumen_dimensiones(
    catalogo
):

    return (
        catalogo
        .groupby(
            "Dimension",
            as_index=False
        )
        .agg(
            Indicadores_Total=(
                "Variable",
                "count"
            ),

            Indicadores_Incluir=(
                "Decision",
                lambda x:
                (
                    x
                    == "INCLUIR"
                ).sum()
            ),

            Indicadores_Control=(
                "Decision",
                lambda x:
                (
                    x
                    == "CONTROL"
                ).sum()
            )
        )
    )


# ============================================================
# CONTROL DE RANGO
# ============================================================

def validar_rangos(
    matriz
):

    filas = []

    for columna in matriz.columns:

        if columna == "Departamento":
            continue

        s = pd.to_numeric(
            matriz[
                columna
            ],
            errors="coerce"
        )

        fuera = int(
            (
                (s < -1e-9)
                |
                (s > 100 + 1e-9)
            ).sum()
        )

        filas.append(
            {
                "Variable_Normalizada":
                    columna,

                "Minimo":
                    s.min(),

                "Maximo":
                    s.max(),

                "Fuera_Rango_0_100":
                    fuera,

                "Faltantes":
                    int(
                        s.isna().sum()
                    )
            }
        )

    return pd.DataFrame(
        filas
    )


# ============================================================
# METODOLOGÍA
# ============================================================

def construir_metodologia():

    return pd.DataFrame(
        {
            "Tema": [
                "Objetivo",
                "Selección",
                "Redundancia",
                "Outliers",
                "Winsorización",
                "Normalización",
                "Sentido",
                "Escala",
                "Variables de control",
                "Pesos",
                "Ranking",
                "Capital natural"
            ],

            "Criterio": [
                (
                    "Seleccionar y normalizar los "
                    "indicadores candidatos del IRNA-C."
                ),

                (
                    "Se selecciona una variable por "
                    "fenómeno cuando existe redundancia."
                ),

                (
                    "Los pares altamente correlacionados "
                    "no ingresan simultáneamente cuando "
                    "representan el mismo fenómeno."
                ),

                (
                    "Se identificó una presencia relevante "
                    "de valores extremos en el diagnóstico 24."
                ),

                (
                    "Cada indicador se limita a sus "
                    "percentiles 5 y 95 antes de normalizar."
                ),

                (
                    "Se utiliza transformación Min-Max "
                    "sobre la serie winsorizada."
                ),

                (
                    "POSITIVO significa que un valor mayor "
                    "favorece la competitividad."
                ),

                (
                    "Todos los indicadores quedan expresados "
                    "en escala 0 a 100."
                ),

                (
                    "Las variables CONTROL permanecen en "
                    "el catálogo pero no entrarán inicialmente "
                    "al índice."
                ),

                (
                    "Esta etapa no asigna pesos."
                ),

                (
                    "Esta etapa no genera ranking IRNA-C."
                ),

                (
                    "La superficie protegida territorial "
                    "continúa siendo referencial hasta contar "
                    "con intersección GIS exacta."
                )
            ]
        }
    )


# ============================================================
# FORMATO EXCEL
# ============================================================

def formatear_excel(ruta):

    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(
        ruta
    )

    for ws in wb.worksheets:

        ws.freeze_panes = "A2"

        if (
            ws.max_row > 1
            and
            ws.max_column > 1
        ):

            ws.auto_filter.ref = (
                ws.dimensions
            )

        for celda in ws[1]:

            celda.font = Font(
                bold=True
            )

            celda.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        for columna in range(
            1,
            ws.max_column + 1
        ):

            letra = get_column_letter(
                columna
            )

            ancho = 0

            for fila in range(
                1,
                min(
                    ws.max_row,
                    150
                ) + 1
            ):

                valor = ws.cell(
                    fila,
                    columna
                ).value

                if valor is not None:

                    ancho = max(
                        ancho,
                        len(
                            str(valor)
                        )
                    )

            ws.column_dimensions[
                letra
            ].width = min(
                max(
                    ancho + 2,
                    12
                ),
                45
            )

    wb.save(
        ruta
    )


# ============================================================
# MAIN
# ============================================================

def main():

    print(
        "=" * 100
    )

    print(
        "RADAR TURISMO NATURALEZA "
        "Y AVENTURA - PERÚ"
    )

    print(
        "25 - SELECCIÓN Y NORMALIZACIÓN IRNA-C"
    )

    print(
        "=" * 100
    )

    df = cargar_datos()

    catalogo = pd.DataFrame(
        INDICADORES
    )

    faltantes = validar_indicadores(
        df,
        catalogo
    )

    (
        normalizados,
        control_normalizacion
    ) = construir_normalizados(
        df,
        catalogo
    )

    matriz_incluidos = (
        construir_matriz_incluidos(
            normalizados,
            catalogo
        )
    )

    dimensiones = (
        construir_resumen_dimensiones(
            catalogo
        )
    )

    control_rangos = validar_rangos(
        matriz_incluidos
    )

    metodologia = (
        construir_metodologia()
    )

    # ========================================================
    # RESUMEN
    # ========================================================

    indicadores_incluir = int(
        (
            catalogo[
                "Decision"
            ]
            == "INCLUIR"
        ).sum()
    )

    indicadores_control = int(
        (
            catalogo[
                "Decision"
            ]
            == "CONTROL"
        ).sum()
    )

    resumen = pd.DataFrame(
        {
            "Indicador": [
                "Territorios",
                "Indicadores definidos",
                "Indicadores incluir",
                "Indicadores control",
                "Variables faltantes",
                "Dimensiones",
                "Percentil winsor inferior",
                "Percentil winsor superior"
            ],

            "Valor": [
                len(
                    df
                ),
                len(
                    catalogo
                ),
                indicadores_incluir,
                indicadores_control,
                len(
                    faltantes
                ),
                catalogo[
                    "Dimension"
                ].nunique(),
                PERCENTIL_INFERIOR,
                PERCENTIL_SUPERIOR
            ]
        }
    )

    # ========================================================
    # EXPORTAR
    # ========================================================

    with pd.ExcelWriter(
        SALIDA,
        engine="openpyxl"
    ) as writer:

        resumen.to_excel(
            writer,
            sheet_name="01_Resumen",
            index=False
        )

        catalogo.to_excel(
            writer,
            sheet_name="02_Catalogo_Indicadores",
            index=False
        )

        matriz_incluidos.to_excel(
            writer,
            sheet_name="03_Normalizados_IRNAC",
            index=False
        )

        normalizados.to_excel(
            writer,
            sheet_name="04_Todos_Normalizados",
            index=False
        )

        control_normalizacion.to_excel(
            writer,
            sheet_name="05_Control_Normalizacion",
            index=False
        )

        dimensiones.to_excel(
            writer,
            sheet_name="06_Dimensiones",
            index=False
        )

        control_rangos.to_excel(
            writer,
            sheet_name="07_Control_Rangos",
            index=False
        )

        metodologia.to_excel(
            writer,
            sheet_name="08_Metodologia",
            index=False
        )

    formatear_excel(
        SALIDA
    )

    # ========================================================
    # CONSOLA
    # ========================================================

    print(
        "\n" + "=" * 100
    )

    print(
        "RESUMEN"
    )

    print(
        "=" * 100
    )

    print(
        resumen.to_string(
            index=False
        )
    )

    print(
        "\n" + "=" * 100
    )

    print(
        "ARQUITECTURA POR DIMENSIÓN"
    )

    print(
        "=" * 100
    )

    print(
        dimensiones.to_string(
            index=False
        )
    )

    print(
        "\n" + "=" * 100
    )

    print(
        "INDICADORES INCLUIDOS"
    )

    print(
        "=" * 100
    )

    print(
        catalogo[
            catalogo[
                "Decision"
            ]
            == "INCLUIR"
        ][
            [
                "Dimension",
                "Variable",
                "Indicador",
                "Sentido"
            ]
        ]
        .to_string(
            index=False
        )
    )

    print(
        "\n" + "=" * 100
    )

    print(
        "CONTROL DE NORMALIZACIÓN"
    )

    print(
        "=" * 100
    )

    print(
        control_normalizacion[
            [
                "Variable",
                "Decision",
                "P05",
                "P95",
                "Territorios_Modificados_Winsor",
                "Faltantes"
            ]
        ]
        .to_string(
            index=False
        )
    )

    problemas_rango = int(
        control_rangos[
            "Fuera_Rango_0_100"
        ].sum()
    )

    faltantes_norm = int(
        control_rangos[
            "Faltantes"
        ].sum()
    )

    print(
        "\n" + "=" * 100
    )

    print(
        "CONTROL FINAL"
    )

    print(
        "=" * 100
    )

    print(
        f"TERRITORIOS                 : "
        f"{len(matriz_incluidos)}"
    )

    print(
        f"INDICADORES IRNA-C          : "
        f"{indicadores_incluir}"
    )

    print(
        f"DIMENSIONES                 : "
        f"{catalogo['Dimension'].nunique()}"
    )

    print(
        f"VARIABLES FALTANTES         : "
        f"{len(faltantes)}"
    )

    print(
        f"VALORES FUERA DE 0-100      : "
        f"{problemas_rango}"
    )

    print(
        f"FALTANTES NORMALIZADOS      : "
        f"{faltantes_norm}"
    )

    if (
        len(
            faltantes
        ) == 0
        and
        problemas_rango == 0
    ):

        print(
            "\n✓ INDICADORES IRNA-C "
            "NORMALIZADOS CORRECTAMENTE"
        )

    else:

        print(
            "\n⚠ REVISAR CONTROLES"
        )

    print(
        "\nARCHIVO GENERADO:"
    )

    print(
        SALIDA
    )

    print(
        "\n✓ ETAPA 25 COMPLETADA"
    )

    print(
        "✓ AÚN NO SE HAN ASIGNADO PESOS"
    )

    print(
        "✓ AÚN NO SE HA GENERADO RANKING"
    )

    print(
        "FIN"
    )


if __name__ == "__main__":
    main()